#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WMS Barcode Scanner Application
Main application file with multi-tab interface
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pyodbc
import json
import pandas as pd
import datetime
import os
import sys
from typing import Dict, List, Optional, Any

# Import database manager and login window
try:
    from database.database_manager import DatabaseManager
    from ui.login_window import LoginWindow
except ImportError:
    print("Error: Required modules not found")
    sys.exit(1)


class WMSScannerApp:
    """Main WMS Scanner Application"""
    
    def __init__(self, root, connection_info: Optional[Dict[str, Any]] = None):
        self.root = root
        self.root.title("WMS EP Asia Group Co., Ltd.")
        # Window properties are now set in main.py
        
        # Initialize database manager with connection info
        self.db = DatabaseManager(connection_info)
        
        # Session variables
        self.current_job_type = tk.StringVar()
        self.current_sub_job_type = tk.StringVar()
        self.notes_var = tk.StringVar()
        self.barcode_entry_var = tk.StringVar()
        self.job_types_data = {}  # Store job types with ID
        self.sub_job_types_data = {}  # Store sub job types with ID
        self.dependencies_vars = {}  # Store dependency checkboxes
        self.current_selected_job = None  # Currently selected job for dependencies
        self.current_selected_sub_job = None  # Currently selected sub job for management
        
        # Initialize UI
        self.setup_ui()
        
        # Test database connection on startup (only if not using login)
        if not connection_info:
            if not self.db.test_connection():
                messagebox.showerror("ข้อผิดพลาด", "ไม่สามารถเชื่อมต่อฐานข้อมูลได้")
        
        # Bind keyboard shortcuts
        self.root.bind('<F11>', self.toggle_fullscreen)
        self.root.bind('<Alt-Return>', self.toggle_fullscreen)
        self.root.bind('<Control-m>', self.toggle_maximize)
        
        # Load initial data
        self.refresh_job_types()
        
        # Load initial data for scanning tab
        self.root.after(500, self.on_main_job_change)  # Initialize sub job combo
        
        # Load initial scanning history
        self.root.after(1000, self.refresh_scanning_history)  # Load initial history data
    
    def setup_ui(self):
        """Setup the main UI with tabs"""
        # Create notebook (tabs)
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create tabs (ไม่รวม database settings tab)
        self.create_scanning_tab()
        self.create_history_tab()        
        self.create_reports_tab()
        self.create_import_tab()
        self.create_settings_tab()
        self.create_sub_job_settings_tab()
    


    def create_settings_tab(self):
        """Create settings tab for job type management"""
        settings_frame = ttk.Frame(self.notebook)
        self.notebook.add(settings_frame, text="จัดการประเภทงานหลัก")

        # Job Types Management
        job_frame = ttk.LabelFrame(settings_frame, text="จัดการประเภทงาน", padding=10)
        job_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Add job type
        add_frame = ttk.Frame(job_frame)
        add_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(add_frame, text="เพิ่มประเภทงาน:").pack(side=tk.LEFT)
        self.new_job_entry = ttk.Entry(add_frame, width=20)
        self.new_job_entry.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(add_frame, text="เพิ่ม", command=self.add_job_type).pack(side=tk.LEFT, padx=5)
        
        # Job types list with dependencies
        list_frame = ttk.Frame(job_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Left side - Job list
        left_frame = ttk.Frame(list_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        
        ttk.Label(left_frame, text="รายการประเภทงาน:").pack(anchor=tk.W)
        
        # Listbox with scrollbar
        job_list_frame = ttk.Frame(left_frame)
        job_list_frame.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(job_list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.job_listbox = tk.Listbox(job_list_frame, yscrollcommand=scrollbar.set, width=25)
        self.job_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.job_listbox.yview)
        self.job_listbox.bind('<<ListboxSelect>>', self.on_job_select)
        
        # Delete button
        ttk.Button(left_frame, text="ลบประเภทงานที่เลือก", 
                  command=self.delete_job_type).pack(pady=5)
        
        # Right side - Dependencies
        right_frame = ttk.Frame(list_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        dependencies_frame = ttk.LabelFrame(right_frame, text="งานที่ต้องทำก่อน (Dependencies)", padding=10)
        dependencies_frame.pack(fill=tk.BOTH, expand=True)
        
        self.selected_job_label = ttk.Label(dependencies_frame, text="เลือกประเภทงานเพื่อตั้งค่า Dependencies", 
                                           font=("Arial", 10, "bold"))
        self.selected_job_label.pack(pady=5)
        
        # Dependencies checkboxes frame
        self.dependencies_frame = ttk.Frame(dependencies_frame)
        self.dependencies_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Save dependencies button
        self.save_dependencies_btn = ttk.Button(dependencies_frame, text="บันทึกการตั้งค่า", 
                                               command=self.save_dependencies, state=tk.DISABLED)
        self.save_dependencies_btn.pack(pady=10)
        
        # Store dependencies data
        self.dependencies_vars = {}
        self.current_selected_job = None
    
    def create_sub_job_settings_tab(self):
        """Create sub job types management tab"""
        sub_job_frame = ttk.Frame(self.notebook)
        self.notebook.add(sub_job_frame, text="จัดการประเภทงานย่อย")
        
        # Main container
        main_container = ttk.Frame(sub_job_frame)
        main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Left side - Main job selection
        left_frame = ttk.LabelFrame(main_container, text="เลือกประเภทงานหลัก", padding=10)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        
        # Main job listbox
        ttk.Label(left_frame, text="ประเภทงานหลัก:").pack(anchor=tk.W)
        
        main_job_list_frame = ttk.Frame(left_frame)
        main_job_list_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        main_scrollbar = ttk.Scrollbar(main_job_list_frame)
        main_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.main_job_listbox = tk.Listbox(main_job_list_frame, yscrollcommand=main_scrollbar.set, width=25)
        self.main_job_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        main_scrollbar.config(command=self.main_job_listbox.yview)
        self.main_job_listbox.bind('<<ListboxSelect>>', self.on_main_job_select_for_sub)
        
        # Right side - Sub job management
        right_frame = ttk.LabelFrame(main_container, text="จัดการประเภทงานย่อย", padding=10)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        # Selected main job display
        self.selected_main_job_label = ttk.Label(right_frame, text="เลือกประเภทงานหลักเพื่อจัดการงานย่อย", 
                                               font=("Arial", 10, "bold"))
        self.selected_main_job_label.pack(pady=5)
        
        # Add new sub job
        add_sub_frame = ttk.Frame(right_frame)
        add_sub_frame.pack(fill=tk.X, pady=10)
        
        ttk.Label(add_sub_frame, text="ชื่อประเภทงานย่อย:").pack(anchor=tk.W)
        self.new_sub_job_entry = ttk.Entry(add_sub_frame, width=30)
        self.new_sub_job_entry.pack(fill=tk.X, pady=2)
        
        ttk.Label(add_sub_frame, text="คำอธิบาย (ไม่บังคับ):").pack(anchor=tk.W, pady=(10,0))
        self.sub_job_desc_entry = ttk.Entry(add_sub_frame, width=30)
        self.sub_job_desc_entry.pack(fill=tk.X, pady=2)
        
        add_sub_btn_frame = ttk.Frame(add_sub_frame)
        add_sub_btn_frame.pack(fill=tk.X, pady=10)
        
        self.add_sub_job_btn = ttk.Button(add_sub_btn_frame, text="เพิ่มประเภทงานย่อย", 
                                         command=self.add_sub_job_type, state=tk.DISABLED)
        self.add_sub_job_btn.pack(side=tk.LEFT)
        
        # Sub job list
        sub_list_frame = ttk.Frame(right_frame)
        sub_list_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        ttk.Label(sub_list_frame, text="รายการประเภทงานย่อย:").pack(anchor=tk.W)
        
        sub_job_list_container = ttk.Frame(sub_list_frame)
        sub_job_list_container.pack(fill=tk.BOTH, expand=True, pady=5)
        
        sub_scrollbar = ttk.Scrollbar(sub_job_list_container)
        sub_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.sub_job_listbox = tk.Listbox(sub_job_list_container, yscrollcommand=sub_scrollbar.set)
        self.sub_job_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sub_scrollbar.config(command=self.sub_job_listbox.yview)
        self.sub_job_listbox.bind('<<ListboxSelect>>', self.on_sub_job_select)
        
        # Sub job management buttons
        sub_btn_frame = ttk.Frame(sub_list_frame)
        sub_btn_frame.pack(fill=tk.X, pady=5)
        
        self.edit_sub_job_btn = ttk.Button(sub_btn_frame, text="แก้ไข", 
                                          command=self.edit_sub_job_type, state=tk.DISABLED)
        self.edit_sub_job_btn.pack(side=tk.LEFT, padx=5)
        
        self.delete_sub_job_btn = ttk.Button(sub_btn_frame, text="ลบ", 
                                            command=self.delete_sub_job_type, state=tk.DISABLED)
        self.delete_sub_job_btn.pack(side=tk.LEFT, padx=5)
        
        # Store sub job data
        self.current_selected_main_job_id = None
        self.current_selected_sub_job_id = None
    
    def create_scanning_tab(self):
        """Create scanning tab with left-right split layout"""
        scan_frame = ttk.Frame(self.notebook)
        self.notebook.add(scan_frame, text="หน้าจอหลัก")
        
        # Create horizontal split layout using PanedWindow
        paned_window = ttk.PanedWindow(scan_frame, orient=tk.HORIZONTAL)
        paned_window.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # LEFT PANEL - Scanning controls and today's summary
        left_panel = ttk.Frame(paned_window)
        paned_window.add(left_panel, weight=1)  # 20% of width
        
        # Main scanning area (left panel)
        main_frame = ttk.LabelFrame(left_panel, text="สแกนบาร์โค้ด", padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Job type selection
        job_frame = ttk.Frame(main_frame)
        job_frame.pack(fill=tk.X, pady=10)
        
        # Main job type
        main_job_frame = ttk.Frame(job_frame)
        main_job_frame.pack(fill=tk.X, pady=5)
        ttk.Label(main_job_frame, text="ประเภทงานหลัก:", font=("Arial", 11), width=12).pack(side=tk.LEFT)
        self.job_combo = ttk.Combobox(main_job_frame, textvariable=self.current_job_type,
                                     state="readonly", font=("Arial", 11), width=18)
        self.job_combo.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)
        self.job_combo.bind('<<ComboboxSelected>>', self.on_main_job_change)
        
        # Sub job type
        sub_job_frame = ttk.Frame(job_frame)
        sub_job_frame.pack(fill=tk.X, pady=5)
        ttk.Label(sub_job_frame, text="ประเภทงานย่อย:", font=("Arial", 11), width=12).pack(side=tk.LEFT)
        self.sub_job_combo = ttk.Combobox(sub_job_frame, textvariable=self.current_sub_job_type,
                                         state="readonly", font=("Arial", 11), width=18)
        self.sub_job_combo.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)
        self.sub_job_combo.bind('<<ComboboxSelected>>', self.on_sub_job_change)
        
        # Notes
        notes_frame = ttk.Frame(job_frame)
        notes_frame.pack(fill=tk.X, pady=5)
        ttk.Label(notes_frame, text="หมายเหตุ:", font=("Arial", 11), width=12).pack(side=tk.LEFT)
        self.notes_entry = ttk.Entry(notes_frame, textvariable=self.notes_var,
                                    font=("Arial", 11))
        self.notes_entry.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)
        
        # Bind notes entry events
        self.notes_entry.bind('<KeyRelease>', self.on_notes_change)
        self.notes_var.trace('w', self.on_notes_var_change)
        
        # Barcode input
        barcode_frame = ttk.Frame(main_frame)
        barcode_frame.pack(fill=tk.X, pady=20)
        
        ttk.Label(barcode_frame, text="บาร์โค้ด:", font=("Arial", 12, "bold")).pack(side=tk.TOP, anchor=tk.W)
        self.barcode_entry = ttk.Entry(barcode_frame, textvariable=self.barcode_entry_var,
                                      font=("Arial", 13))
        self.barcode_entry.pack(fill=tk.X, pady=5)
        self.barcode_entry.bind('<Return>', self.process_barcode)
        
        # Today Summary section (left panel)
        self.today_summary_frame = ttk.LabelFrame(main_frame, text="📊 สรุปงานวันนี้", padding=10)
        self.today_summary_frame.pack(fill=tk.X, pady=10)
        
        # Summary content frame
        summary_content_frame = ttk.Frame(self.today_summary_frame)
        summary_content_frame.pack(fill=tk.X)
        
        # Summary labels dictionary for easy updates
        self.summary_labels = {}
        
        # Job Type row
        job_row = ttk.Frame(summary_content_frame)
        job_row.pack(fill=tk.X, pady=2)
        ttk.Label(job_row, text="งานหลัก:", font=("Arial", 10, "bold"), foreground="#0c5184").pack(side=tk.LEFT)
        self.summary_labels['job_type'] = ttk.Label(job_row, text="ไม่ได้เลือก", font=("Arial", 10))
        self.summary_labels['job_type'].pack(side=tk.LEFT, padx=(10, 0))
        
        # Sub Job Type row
        sub_job_row = ttk.Frame(summary_content_frame)
        sub_job_row.pack(fill=tk.X, pady=2)
        ttk.Label(sub_job_row, text="งานรอง:", font=("Arial", 10, "bold"), foreground="#0c5184").pack(side=tk.LEFT)
        self.summary_labels['sub_job_type'] = ttk.Label(sub_job_row, text="ไม่ได้เลือก", font=("Arial", 10))
        self.summary_labels['sub_job_type'].pack(side=tk.LEFT, padx=(10, 0))
        
        # Notes Filter row
        notes_filter_row = ttk.Frame(summary_content_frame)
        notes_filter_row.pack(fill=tk.X, pady=2)
        ttk.Label(notes_filter_row, text="กรองหมายเหตุ:", font=("Arial", 10, "bold"), foreground="#0c5184").pack(side=tk.LEFT)
        self.summary_labels['notes_filter'] = ttk.Label(notes_filter_row, text="ไม่มี", font=("Arial", 10))
        self.summary_labels['notes_filter'].pack(side=tk.LEFT, padx=(10, 0))
        
        # Count row (prominent display)
        count_row = ttk.Frame(summary_content_frame)
        count_row.pack(fill=tk.X, pady=(10, 5))
        ttk.Label(count_row, text="จำนวนที่สแกนวันนี้:", font=("Arial", 11, "bold"), foreground="#0c5184").pack(anchor=tk.W)
        
        # Count frame with background color
        count_frame = ttk.Frame(count_row)
        count_frame.pack(pady=5)
        self.summary_labels['count'] = ttk.Label(count_frame, text="0", 
                                                font=("Arial", 24, "bold"), 
                                                foreground="#0c5184",
                                                # background="#f0f8ff",
                                                padding=5)
        self.summary_labels['count'].pack()
        
        # Last update row
        update_row = ttk.Frame(summary_content_frame)
        update_row.pack(fill=tk.X, pady=2)
        ttk.Label(update_row, text="อัปเดทล่าสุด:", font=("Arial", 9), foreground="gray").pack(side=tk.LEFT)
        self.summary_labels['last_update'] = ttk.Label(update_row, text="ยังไม่มีข้อมูล", font=("Arial", 9), foreground="gray")
        self.summary_labels['last_update'].pack(side=tk.LEFT, padx=(5, 0))
        
        # RIGHT PANEL - Recent scan history
        right_panel = ttk.Frame(paned_window)
        paned_window.add(right_panel, weight=4)  # 80% of width
        
        # History table for scanning tab (right panel)
        history_scan_frame = ttk.LabelFrame(right_panel, text="ประวัติการสแกนล่าสุด", padding=10)
        history_scan_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create history table for scanning tab
        self.create_scanning_history_table(history_scan_frame)
        
        # Focus on barcode entry when tab is selected
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)
    
    def create_import_tab(self):
        """Create import data tab"""
        import_frame = ttk.Frame(self.notebook)
        self.notebook.add(import_frame, text="นำเข้าข้อมูล")
        
        # Main container
        main_container = ttk.Frame(import_frame)
        main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Template section
        template_frame = ttk.LabelFrame(main_container, text="เทมเพลต Excel", padding=10)
        template_frame.pack(fill=tk.X, pady=(0, 10))
        
        template_info = ttk.Label(template_frame, 
                                 text="ดาวน์โหลดเทมเพลต Excel เพื่อเตรียมข้อมูลสำหรับการนำเข้า\nคอลัมน์ที่ต้องมี: บาร์โค้ด, ID_ประเภทงานหลัก, ID_ประเภทงานย่อย, หมายเหตุ (ไม่บังคับ)\nหมายเหตุ: ใช้ ID แทนชื่อสำหรับประเภทงาน (คอลัมน์สีแดง = บังคับ, สีฟ้า = ไม่บังคับ)",
                                 justify=tk.LEFT)
        template_info.pack(pady=5)
        
        template_btn_frame = ttk.Frame(template_frame)
        template_btn_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(template_btn_frame, text="ดาวน์โหลดเทมเพลต Excel", 
                  command=self.download_template).pack(side=tk.LEFT, padx=5)
        ttk.Button(template_btn_frame, text="ดูตัวอย่างข้อมูล", 
                  command=self.show_sample_data).pack(side=tk.LEFT, padx=5)
        
        # Import section
        import_section_frame = ttk.LabelFrame(main_container, text="นำเข้าข้อมูล", padding=10)
        import_section_frame.pack(fill=tk.BOTH, expand=True)
        
        # File selection
        file_frame = ttk.Frame(import_section_frame)
        file_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(file_frame, text="เลือกไฟล์ Excel:").pack(side=tk.LEFT)
        self.import_file_var = tk.StringVar()
        self.import_file_entry = ttk.Entry(file_frame, textvariable=self.import_file_var, width=50, state="readonly")
        self.import_file_entry.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)
        
        ttk.Button(file_frame, text="เลือกไฟล์", command=self.select_import_file).pack(side=tk.LEFT, padx=5)
        
        # Import options
        options_frame = ttk.Frame(import_section_frame)
        options_frame.pack(fill=tk.X, pady=10)
        
        # Skip duplicates option
        self.skip_duplicates_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="ข้ามข้อมูลที่ซ้ำกัน (แนะนำ)", 
                       variable=self.skip_duplicates_var).pack(anchor=tk.W)
        
        # Validate dependencies option
        self.validate_dependencies_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="ตรวจสอบ Dependencies ก่อนนำเข้า", 
                       variable=self.validate_dependencies_var).pack(anchor=tk.W)
        
        # Import buttons
        import_btn_frame = ttk.Frame(import_section_frame)
        import_btn_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(import_btn_frame, text="ตรวจสอบข้อมูล", 
                  command=self.validate_import_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(import_btn_frame, text="นำเข้าข้อมูล", 
                  command=self.import_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(import_btn_frame, text="ล้างข้อมูล", 
                  command=self.clear_import_data).pack(side=tk.LEFT, padx=5)
        
        # Preview/Results section
        preview_frame = ttk.LabelFrame(import_section_frame, text="ตรวจสอบข้อมูลก่อนนำเข้า", padding=10)
        preview_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Create preview table
        self.create_import_preview_table(preview_frame)
        
        # Status
        self.import_status_label = ttk.Label(import_section_frame, text="เลือกไฟล์ Excel เพื่อเริ่มต้น", 
                                           foreground="blue")
        self.import_status_label.pack(pady=5)
        
        # Store import data
        self.import_data_df = None
        self.import_validation_results = None
    
    def create_history_tab(self):
        """Create history viewing tab"""
        history_frame = ttk.Frame(self.notebook)
        self.notebook.add(history_frame, text="ประวัติ")
        
        # Search controls
        search_frame = ttk.LabelFrame(history_frame, text="ค้นหา", padding=10)
        search_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # Date search
        date_frame = ttk.Frame(search_frame)
        date_frame.pack(fill=tk.X, pady=2)
        ttk.Label(date_frame, text="วันที่:").pack(side=tk.LEFT)
        self.date_entry = ttk.Entry(date_frame, width=12)
        self.date_entry.pack(side=tk.LEFT, padx=5)
        self.date_entry.insert(0, datetime.date.today().strftime("%Y-%m-%d"))
        self.date_entry.bind('<Return>', self.search_history)
        
        # Barcode search
        barcode_search_frame = ttk.Frame(search_frame)
        barcode_search_frame.pack(fill=tk.X, pady=2)
        ttk.Label(barcode_search_frame, text="บาร์โค้ด:").pack(side=tk.LEFT)
        self.search_barcode_entry = ttk.Entry(barcode_search_frame, width=30)
        self.search_barcode_entry.pack(side=tk.LEFT, padx=5)
        self.search_barcode_entry.bind('<Return>', self.search_history)
        
        # Job type search
        job_search_frame = ttk.Frame(search_frame)
        job_search_frame.pack(fill=tk.X, pady=2)
        ttk.Label(job_search_frame, text="ประเภทงาน:").pack(side=tk.LEFT)
        self.search_job_combo = ttk.Combobox(job_search_frame, state="readonly", width=25)
        self.search_job_combo.pack(side=tk.LEFT, padx=5)
        self.search_job_combo.bind('<<ComboboxSelected>>', self.search_history)
        
        # Results table
        self.create_history_table(history_frame)
    
    def create_history_table(self, parent):
        """Create table for history display"""
        table_frame = ttk.Frame(parent)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Treeview for table display
        columns = ('ID', 'บาร์โค้ด', 'วันที่/เวลา', 'ประเภทงานหลัก', 'ประเภทงานย่อย', 'หมายเหตุ', 'ผู้ใช้')
        self.history_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=20)
        
        # Define headings
        for col in columns:
            self.history_tree.heading(col, text=col)
            if col == 'หมายเหตุ':
                self.history_tree.column(col, width=200)
            else:
                self.history_tree.column(col, width=150)
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.history_tree.yview)
        h_scrollbar = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.history_tree.xview)
        self.history_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Pack table and scrollbars
        self.history_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
    
    def create_reports_tab(self):
        """Create reports tab with dynamic filtering like Web App"""
        reports_frame = ttk.Frame(self.notebook)
        self.notebook.add(reports_frame, text="รายงาน")
        
        # Filter selection frame
        filter_frame = ttk.LabelFrame(reports_frame, text="เลือกเงื่อนไขรายงาน", padding=10)
        filter_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # Date selection
        date_frame = ttk.Frame(filter_frame)
        date_frame.pack(fill=tk.X, pady=5)
        ttk.Label(date_frame, text="วันที่:").pack(side=tk.LEFT)
        self.report_date_var = tk.StringVar(value=datetime.date.today().strftime("%Y-%m-%d"))
        report_date_entry = ttk.Entry(date_frame, textvariable=self.report_date_var, width=15)
        report_date_entry.pack(side=tk.LEFT, padx=10)
        
        # Job Type selection
        job_type_frame = ttk.Frame(filter_frame)
        job_type_frame.pack(fill=tk.X, pady=5)
        ttk.Label(job_type_frame, text="งานหลัก:").pack(side=tk.LEFT)
        self.report_job_type_var = tk.StringVar()
        self.report_job_type_combo = ttk.Combobox(job_type_frame, textvariable=self.report_job_type_var, 
                                                  state="readonly", width=30)
        self.report_job_type_combo.pack(side=tk.LEFT, padx=10)
        self.report_job_type_combo.bind("<<ComboboxSelected>>", self.on_report_job_type_change)
        
        # Sub Job Type selection
        sub_job_type_frame = ttk.Frame(filter_frame)
        sub_job_type_frame.pack(fill=tk.X, pady=5)
        ttk.Label(sub_job_type_frame, text="งานรอง:").pack(side=tk.LEFT)
        self.report_sub_job_type_var = tk.StringVar()
        self.report_sub_job_type_combo = ttk.Combobox(sub_job_type_frame, textvariable=self.report_sub_job_type_var, 
                                                      state="readonly", width=30)
        self.report_sub_job_type_combo.pack(side=tk.LEFT, padx=10)
        
        # Note filter
        note_filter_frame = ttk.Frame(filter_frame)
        note_filter_frame.pack(fill=tk.X, pady=5)
        ttk.Label(note_filter_frame, text="กรองหมายเหตุ:").pack(side=tk.LEFT)
        self.report_note_filter_var = tk.StringVar()
        note_filter_entry = ttk.Entry(note_filter_frame, textvariable=self.report_note_filter_var, width=40)
        note_filter_entry.pack(side=tk.LEFT, padx=10)
        
        # Run buttons
        button_frame = ttk.Frame(filter_frame)
        button_frame.pack(fill=tk.X, pady=10)
        ttk.Button(button_frame, text="📊 ดูรายงาน", command=self.run_report).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="ส่งออก Excel", command=self.export_report).pack(side=tk.LEFT, padx=5)
        
        # Results table
        self.create_report_table(reports_frame)
        
        # Load available job types
        self.refresh_report_job_types()
    
    def create_report_table(self, parent):
        """Create table for report display"""
        table_frame = ttk.Frame(parent)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Dynamic treeview (columns will be set when data loads)
        self.report_tree = ttk.Treeview(table_frame, show='headings', height=25)
        
        # Scrollbars
        v_scrollbar_report = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.report_tree.yview)
        h_scrollbar_report = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.report_tree.xview)
        self.report_tree.configure(yscrollcommand=v_scrollbar_report.set, xscrollcommand=h_scrollbar_report.set)
        
        # Pack
        self.report_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        v_scrollbar_report.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar_report.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Store report data for export
        self.current_report_data = []
    
    def create_scanning_history_table(self, parent):
        """Create history table for scanning tab"""
        table_frame = ttk.Frame(parent)
        table_frame.pack(fill=tk.BOTH, expand=True)
        
        # Treeview for table display
        columns = ('ID', 'บาร์โค้ด', 'วันที่/เวลา', 'ประเภทงานหลัก', 'ประเภทงานย่อย', 'หมายเหตุ', 'ผู้ใช้')
        self.scan_history_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)
        
        # Define headings with optimized widths for 80% right panel
        column_widths = {
            'ID': 80,
            'บาร์โค้ด': 150,
            'วันที่/เวลา': 180,
            'ประเภทงานหลัก': 150,
            'ประเภทงานย่อย': 150,
            'หมายเหตุ': 200,
            'ผู้ใช้': 120
        }
        
        for col in columns:
            self.scan_history_tree.heading(col, text=col)
            self.scan_history_tree.column(col, width=column_widths.get(col, 100), minwidth=50)
        
        # Scrollbars
        v_scrollbar_scan = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.scan_history_tree.yview)
        h_scrollbar_scan = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.scan_history_tree.xview)
        self.scan_history_tree.configure(yscrollcommand=v_scrollbar_scan.set, xscrollcommand=h_scrollbar_scan.set)
        
        # Pack table and scrollbars
        self.scan_history_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        v_scrollbar_scan.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar_scan.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Bind right-click context menu
        self.scan_history_tree.bind("<Button-3>", self.show_scan_context_menu)  # Right-click
        
        # Create context menu
        self.scan_context_menu = tk.Menu(self.root, tearoff=0)
        self.scan_context_menu.add_command(label="แก้ไขข้อมูล", command=self.edit_scan_record)
        self.scan_context_menu.add_separator()
        self.scan_context_menu.add_command(label="ลบข้อมูล", command=self.delete_scan_record)
        self.scan_context_menu.add_separator()
        self.scan_context_menu.add_command(label="ยกเลิก", command=lambda: self.scan_context_menu.unpost())
    
    def create_import_preview_table(self, parent):
        """Create table for import data preview"""
        table_frame = ttk.Frame(parent)
        table_frame.pack(fill=tk.BOTH, expand=True)
        
        # Treeview for table display
        columns = ('สถานะ', 'บาร์โค้ด', 'ประเภทงานหลัก', 'ประเภทงานย่อย', 'หมายเหตุ', 'ข้อผิดพลาด')
        self.import_preview_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)
        
        # Define headings
        for col in columns:
            self.import_preview_tree.heading(col, text=col)
            if col == 'สถานะ':
                self.import_preview_tree.column(col, width=80)
            elif col == 'ข้อผิดพลาด':
                self.import_preview_tree.column(col, width=250)
            elif col == 'หมายเหตุ':
                self.import_preview_tree.column(col, width=150)
            else:
                self.import_preview_tree.column(col, width=120)
        
        # Scrollbars
        v_scrollbar_import = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.import_preview_tree.yview)
        h_scrollbar_import = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.import_preview_tree.xview)
        self.import_preview_tree.configure(yscrollcommand=v_scrollbar_import.set, xscrollcommand=h_scrollbar_import.set)
        
        # Pack table and scrollbars
        self.import_preview_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        v_scrollbar_import.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar_import.pack(side=tk.BOTTOM, fill=tk.X)
    
    # Event handlers and business logic methods
    def on_tab_changed(self, event):
        """Handle tab change events"""
        selected_tab = event.widget.tab('current')['text']
        if selected_tab == "หน้าจอหลัก":
            self.root.after(100, lambda: self.barcode_entry.focus_set())
            # Refresh scanning history when entering scan tab
            self.refresh_scanning_history()
    
    def refresh_scanning_history(self):
        """Refresh scanning history table with recent data filtered by user selections"""
        try:
            print("Refreshing scanning history...")  # Debug message
            
            # Get current user selections for filtering
            job_type_name = self.current_job_type.get()
            sub_job_type_name = self.current_sub_job_type.get()
            notes_filter = self.notes_var.get().strip()
            
            # Base query with joins
            query = """
                SELECT TOP 50 
                    sl.id, 
                    sl.barcode, 
                    sl.scan_date, 
                    sl.job_type as main_job_name,
                    sjt.sub_job_name,
                    sl.notes,
                    sl.user_id 
                FROM scan_logs sl
                LEFT JOIN sub_job_types sjt ON sl.sub_job_id = sjt.id
            """
            
            # Build WHERE conditions based on user selections
            conditions = []
            params = []
            
            # Filter by job type if selected
            if job_type_name:
                # Get job type ID
                job_result = self.db.execute_query("SELECT id FROM job_types WHERE job_name = ?", (job_type_name,))
                if job_result:
                    job_type_id = job_result[0]['id']
                    conditions.append("sl.job_id = ?")
                    params.append(job_type_id)
                    
                    # Filter by sub job type if selected
                    if sub_job_type_name:
                        sub_result = self.db.execute_query(
                            "SELECT id FROM sub_job_types WHERE sub_job_name = ? AND main_job_id = ?", 
                            (sub_job_type_name, job_type_id)
                        )
                        if sub_result:
                            sub_job_id = sub_result[0]['id']
                            conditions.append("sl.sub_job_id = ?")
                            params.append(sub_job_id)
            
            # Filter by notes if provided
            if notes_filter:
                conditions.append("sl.notes LIKE ?")
                params.append(f"%{notes_filter}%")
            
            # Add WHERE clause if there are conditions
            if conditions:
                query += " WHERE " + " AND ".join(conditions)
            
            # Add ORDER BY clause
            query += " ORDER BY sl.scan_date DESC"
            
            results = self.db.execute_query(query, tuple(params))
            
            print(f"Found {len(results)} records")  # Debug message
            
            # Clear existing data
            for item in self.scan_history_tree.get_children():
                self.scan_history_tree.delete(item)
            
            # Populate table
            for row in results:
                scan_date = row['scan_date'].strftime("%Y-%m-%d %H:%M:%S") if row['scan_date'] else ""
                sub_job_name = row['sub_job_name'] if row['sub_job_name'] else ""
                notes = row['notes'] if row['notes'] else ""
                
                self.scan_history_tree.insert('', tk.END, values=(
                    row['id'],
                    row['barcode'],
                    scan_date,
                    row['main_job_name'],
                    sub_job_name,
                    notes,
                    row['user_id']
                ))
            
            print("Scanning history refreshed successfully")  # Debug message
                
        except Exception as e:
            # Don't show error message for history refresh to avoid interrupting scanning
            print(f"Error refreshing scanning history: {str(e)}")
            print(f"Error details: {type(e).__name__}: {str(e)}")
    
    def refresh_job_types(self):
        """Refresh job types from database"""
        try:
            results = self.db.execute_query("SELECT id, job_name FROM job_types ORDER BY job_name")
            job_names = [row['job_name'] for row in results]
            
            # Store job types with ID for validation
            self.job_types_data = {row['job_name']: row['id'] for row in results}
            
            # Update UI components
            self.job_combo['values'] = job_names
            self.search_job_combo['values'] = [''] + job_names  # Include empty option for search
            
            # Update settings listbox
            self.job_listbox.delete(0, tk.END)
            for result in results:
                display_text = f"{result['job_name']} (ID: {result['id']})"
                self.job_listbox.insert(tk.END, display_text)
            
            # Refresh dependencies display
            self.refresh_dependencies_display()
            
            # Refresh main job list for sub job management
            if hasattr(self, 'main_job_listbox'):
                self.refresh_main_job_list_for_sub()
                
        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถโหลดประเภทงานได้: {str(e)}")
    
    def refresh_main_job_list_for_sub(self):
        """Refresh main job list in sub job management tab"""
        try:
            # Clear existing items
            self.main_job_listbox.delete(0, tk.END)
            
            # Add job types to main job listbox
            for job_name, job_id in self.job_types_data.items():
                display_text = f"{job_name} (ID: {job_id})"
                self.main_job_listbox.insert(tk.END, display_text)
                
        except Exception as e:
            print(f"Error refreshing main job list: {str(e)}")
    
    def on_main_job_select_for_sub(self, event):
        """Handle main job selection for sub job management"""
        selection = self.main_job_listbox.curselection()
        if not selection:
            return
        
        # Get selected job
        display_text = self.main_job_listbox.get(selection[0])
        job_name = display_text.split(" (ID: ")[0]
        job_id = self.job_types_data.get(job_name)
        
        self.current_selected_main_job_id = job_id
        self.selected_main_job_label.config(text=f"จัดการประเภทงานย่อยสำหรับ: {job_name}")
        self.add_sub_job_btn.config(state=tk.NORMAL)
        
        # Load sub job types for this main job
        self.refresh_sub_job_list()
    
    def refresh_sub_job_list(self):
        """Refresh sub job types list for selected main job"""
        if not self.current_selected_main_job_id:
            return
        
        try:
            # Clear existing items
            self.sub_job_listbox.delete(0, tk.END)
            
            # Get sub job types for current main job
            query = """
                SELECT id, sub_job_name, description 
                FROM sub_job_types 
                WHERE main_job_id = ? AND is_active = 1
                ORDER BY sub_job_name
            """
            results = self.db.execute_query(query, (self.current_selected_main_job_id,))
            
            # Update sub_job_types_data
            self.sub_job_types_data = {}
            
            for row in results:
                sub_job_id = row['id']
                sub_job_name = row['sub_job_name']
                description = row['description'] or ""
                
                self.sub_job_types_data[sub_job_name] = sub_job_id
                
                # Display in listbox
                display_text = f"{sub_job_name}"
                if description:
                    display_text += f" - {description}"
                display_text += f" (ID: {sub_job_id})"
                
                self.sub_job_listbox.insert(tk.END, display_text)
                
        except Exception as e:
            print(f"Error refreshing sub job list: {str(e)}")
    
    def on_sub_job_select(self, event):
        """Handle sub job selection"""
        selection = self.sub_job_listbox.curselection()
        if selection:
            self.edit_sub_job_btn.config(state=tk.NORMAL)
            self.delete_sub_job_btn.config(state=tk.NORMAL)
            
            # Get selected sub job ID
            display_text = self.sub_job_listbox.get(selection[0])
            # Extract ID from the end of the string
            if "(ID: " in display_text:
                id_part = display_text.split("(ID: ")[-1].rstrip(")")
                try:
                    self.current_selected_sub_job_id = int(id_part)
                except ValueError:
                    self.current_selected_sub_job_id = None
        else:
            self.edit_sub_job_btn.config(state=tk.DISABLED)
            self.delete_sub_job_btn.config(state=tk.DISABLED)
            self.current_selected_sub_job_id = None
    
    def add_sub_job_type(self):
        """Add new sub job type"""
        if not self.current_selected_main_job_id:
            messagebox.showwarning("คำเตือน", "กรุณาเลือกประเภทงานหลักก่อน")
            return
        
        sub_job_name = self.new_sub_job_entry.get().strip()
        description = self.sub_job_desc_entry.get().strip()
        
        if not sub_job_name:
            messagebox.showwarning("คำเตือน", "กรุณาใส่ชื่อประเภทงานย่อย")
            return
        
        try:
            # Check for duplicate name within the same main job
            check_query = """
                SELECT COUNT(*) as count 
                FROM sub_job_types 
                WHERE main_job_id = ? AND sub_job_name = ? AND is_active = 1
            """
            result = self.db.execute_query(check_query, (self.current_selected_main_job_id, sub_job_name))
            
            if result[0]['count'] > 0:
                messagebox.showwarning("คำเตือน", "ชื่อประเภทงานย่อยนี้มีอยู่แล้ว")
                return
            
            # Insert new sub job type
            insert_query = """
                INSERT INTO sub_job_types (main_job_id, sub_job_name, description) 
                VALUES (?, ?, ?)
            """
            self.db.execute_non_query(insert_query, (self.current_selected_main_job_id, sub_job_name, description))
            
            # Clear input fields
            self.new_sub_job_entry.delete(0, tk.END)
            self.sub_job_desc_entry.delete(0, tk.END)
            
            # Refresh list
            self.refresh_sub_job_list()
            
            messagebox.showinfo("สำเร็จ", f"เพิ่มประเภทงานย่อย '{sub_job_name}' แล้ว")
            
        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถเพิ่มประเภทงานย่อยได้: {str(e)}")
    
    def edit_sub_job_type(self):
        """Edit selected sub job type"""
        if not self.current_selected_sub_job_id:
            messagebox.showwarning("คำเตือน", "กรุณาเลือกประเภทงานย่อยที่จะแก้ไข")
            return
        
        # Get current data
        try:
            query = "SELECT sub_job_name, description FROM sub_job_types WHERE id = ?"
            result = self.db.execute_query(query, (self.current_selected_sub_job_id,))
            
            if not result:
                messagebox.showerror("ข้อผิดพลาด", "ไม่พบข้อมูลประเภทงานย่อย")
                return
            
            current_data = result[0]
            
            # Show edit dialog
            self.show_sub_job_edit_dialog(self.current_selected_sub_job_id, current_data)
            
        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถโหลดข้อมูลได้: {str(e)}")
    
    def delete_sub_job_type(self):
        """Delete selected sub job type"""
        if not self.current_selected_sub_job_id:
            messagebox.showwarning("คำเตือน", "กรุณาเลือกประเภทงานย่อยที่จะลบ")
            return
        
        # Get sub job name for confirmation
        try:
            query = "SELECT sub_job_name FROM sub_job_types WHERE id = ?"
            result = self.db.execute_query(query, (self.current_selected_sub_job_id,))
            
            if not result:
                messagebox.showerror("ข้อผิดพลาด", "ไม่พบข้อมูลประเภทงานย่อย")
                return
            
            sub_job_name = result[0]['sub_job_name']
            
            # Confirm deletion
            if messagebox.askyesno("ยืนยัน", 
                                  f"ต้องการลบประเภทงานย่อย '{sub_job_name}' หรือไม่?\n\n" +
                                  "หมายเหตุ: ข้อมูลการสแกนที่ใช้ประเภทงานย่อยนี้จะยังคงอยู่ แต่จะไม่สามารถเลือกใช้ได้อีก"):
                
                # Soft delete (set is_active = 0)
                update_query = "UPDATE sub_job_types SET is_active = 0 WHERE id = ?"
                self.db.execute_non_query(update_query, (self.current_selected_sub_job_id,))
                
                messagebox.showinfo("สำเร็จ", f"ลบประเภทงานย่อย '{sub_job_name}' แล้ว")
                
                # Refresh list
                self.refresh_sub_job_list()
                
                # Reset button states
                self.edit_sub_job_btn.config(state=tk.DISABLED)
                self.delete_sub_job_btn.config(state=tk.DISABLED)
                self.current_selected_sub_job_id = None
                
        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถลบประเภทงานย่อยได้: {str(e)}")
    
    def show_sub_job_edit_dialog(self, sub_job_id, current_data):
        """Show dialog for editing sub job type"""
        dialog = tk.Toplevel(self.root)
        dialog.title("แก้ไขประเภทงานย่อย")
        dialog.geometry("400x220")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.geometry("+%d+%d" % (self.root.winfo_rootx() + 100, self.root.winfo_rooty() + 100))
        
        # Main frame
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Sub job name
        name_frame = ttk.Frame(main_frame)
        name_frame.pack(fill=tk.X, pady=5)
        ttk.Label(name_frame, text="ชื่อประเภทงานย่อย:").pack(anchor=tk.W)
        name_entry = ttk.Entry(name_frame, width=40)
        name_entry.pack(fill=tk.X, pady=2)
        name_entry.insert(0, current_data['sub_job_name'])
        
        # Description
        desc_frame = ttk.Frame(main_frame)
        desc_frame.pack(fill=tk.X, pady=5)
        ttk.Label(desc_frame, text="คำอธิบาย:").pack(anchor=tk.W)
        desc_entry = ttk.Entry(desc_frame, width=40)
        desc_entry.pack(fill=tk.X, pady=2)
        desc_entry.insert(0, current_data['description'] or "")
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=20)
        
        def save_changes():
            new_name = name_entry.get().strip()
            new_desc = desc_entry.get().strip()
            
            if not new_name:
                messagebox.showwarning("คำเตือน", "กรุณาใส่ชื่อประเภทงานย่อย")
                return
            
            try:
                # Check for duplicate name (excluding current record)
                check_query = """
                    SELECT COUNT(*) as count 
                    FROM sub_job_types 
                    WHERE main_job_id = ? AND sub_job_name = ? AND id != ? AND is_active = 1
                """
                result = self.db.execute_query(check_query, (self.current_selected_main_job_id, new_name, sub_job_id))
                
                if result[0]['count'] > 0:
                    messagebox.showwarning("คำเตือน", "ชื่อประเภทงานย่อยนี้มีอยู่แล้ว")
                    return
                
                # Update database
                update_query = """
                    UPDATE sub_job_types 
                    SET sub_job_name = ?, description = ?, updated_date = GETDATE() 
                    WHERE id = ?
                """
                self.db.execute_non_query(update_query, (new_name, new_desc, sub_job_id))
                
                messagebox.showinfo("สำเร็จ", "อัพเดทข้อมูลเรียบร้อยแล้ว")
                dialog.destroy()
                
                # Refresh list
                self.refresh_sub_job_list()
                
            except Exception as e:
                messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถอัพเดทข้อมูลได้: {str(e)}")
        
        def cancel_edit():
            dialog.destroy()
        
        ttk.Button(button_frame, text="บันทึกการเปลี่ยนแปลง", command=save_changes).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="ยกเลิก", command=cancel_edit).pack(side=tk.LEFT, padx=5)
        
        # Focus on name entry
        name_entry.focus_set()
        name_entry.select_range(0, tk.END)
    
    def on_main_job_change(self, event=None):
        """Handle main job type change in scanning tab"""
        main_job_name = self.current_job_type.get()
        
        if not main_job_name:
            self.sub_job_combo['values'] = []
            self.current_sub_job_type.set('')
            return
        
        # Get main job ID
        main_job_id = self.job_types_data.get(main_job_name)
        if not main_job_id:
            return
        
        try:
            # Load sub job types for selected main job
            query = """
                SELECT sub_job_name 
                FROM sub_job_types 
                WHERE main_job_id = ? AND is_active = 1
                ORDER BY sub_job_name
            """
            results = self.db.execute_query(query, (main_job_id,))
            
            # Update sub job combo
            sub_job_names = [row['sub_job_name'] for row in results]
            self.sub_job_combo['values'] = sub_job_names
            
            # Clear current selection
            self.current_sub_job_type.set('')
            
            # Update today summary
            self.load_today_summary()
            
            # Refresh scanning history with new filter
            self.refresh_scanning_history()
            
        except Exception as e:
            print(f"Error loading sub job types: {str(e)}")
            self.sub_job_combo['values'] = []
            self.current_sub_job_type.set('')
            
            # Clear summary on error
            self.clear_summary()
            # Refresh history to show unfiltered results
            self.refresh_scanning_history()
    
    def on_sub_job_change(self, event=None):
        """Handle sub job type change"""
        try:
            # Update today summary when sub job type changes
            self.load_today_summary()
            
            # Refresh scanning history with new filter
            self.refresh_scanning_history()
        except Exception as e:
            print(f"Error in on_sub_job_change: {str(e)}")
    
    def on_notes_change(self, event=None):
        """Handle notes entry change with debouncing"""
        try:
            # Cancel previous scheduled update
            if hasattr(self, '_notes_update_job'):
                self.root.after_cancel(self._notes_update_job)
            
            # Schedule update after 500ms delay (debouncing)
            def update_all():
                self.load_today_summary()
                self.refresh_scanning_history()
            
            self._notes_update_job = self.root.after(500, update_all)
        except Exception as e:
            print(f"Error in on_notes_change: {str(e)}")
    
    def on_notes_var_change(self, *args):
        """Handle notes variable change"""
        try:
            # Check if UI is ready
            if hasattr(self, 'summary_labels') and self.summary_labels:
                # Update summary when notes variable changes
                self.load_today_summary()
                
                # Refresh scanning history with new filter
                self.refresh_scanning_history()
        except Exception as e:
            print(f"Error in on_notes_var_change: {str(e)}")
    
    def on_job_select(self, event):
        """Handle job selection in listbox"""
        selection = self.job_listbox.curselection()
        if not selection:
            return
        
        # Get selected job name
        display_text = self.job_listbox.get(selection[0])
        job_name = display_text.split(" (ID: ")[0]
        job_id = self.job_types_data.get(job_name)
        
        self.current_selected_job = job_id
        self.selected_job_label.config(text=f"ตั้งค่า Dependencies สำหรับ: {job_name}")
        self.save_dependencies_btn.config(state=tk.NORMAL)
        
        # Refresh dependencies checkboxes
        self.refresh_dependencies_display()
    
    def refresh_dependencies_display(self):
        """Refresh dependencies checkboxes"""
        # Clear existing checkboxes
        for widget in self.dependencies_frame.winfo_children():
            widget.destroy()
        
        self.dependencies_vars.clear()
        
        if not self.current_selected_job:
            return
        
        try:
            # Get all job types except current selected
            all_jobs = self.db.execute_query("SELECT id, job_name FROM job_types WHERE id != ? ORDER BY job_name", 
                                           (self.current_selected_job,))
            
            # Get current dependencies
            current_deps = self.db.execute_query(
                "SELECT required_job_id FROM job_dependencies WHERE job_id = ?", 
                (self.current_selected_job,)
            )
            required_ids = [row['required_job_id'] for row in current_deps]
            
            # Create checkboxes
            for job in all_jobs:
                var = tk.BooleanVar()
                var.set(job['id'] in required_ids)
                self.dependencies_vars[job['id']] = var
                
                checkbox = ttk.Checkbutton(
                    self.dependencies_frame, 
                    text=f"{job['job_name']} (ID: {job['id']})",
                    variable=var
                )
                checkbox.pack(anchor=tk.W, pady=2)
                
        except Exception as e:
            print(f"Error refreshing dependencies: {str(e)}")
    
    def save_dependencies(self):
        """Save dependencies for selected job"""
        if not self.current_selected_job:
            return
        
        try:
            # Delete existing dependencies
            self.db.execute_non_query("DELETE FROM job_dependencies WHERE job_id = ?", 
                                     (self.current_selected_job,))
            
            # Insert new dependencies
            for job_id, var in self.dependencies_vars.items():
                if var.get():
                    self.db.execute_non_query(
                        "INSERT INTO job_dependencies (job_id, required_job_id) VALUES (?, ?)",
                        (self.current_selected_job, job_id)
                    )
            
            messagebox.showinfo("สำเร็จ", "บันทึกการตั้งค่า Dependencies แล้ว")
            
        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถบันทึกการตั้งค่าได้: {str(e)}")
    
    def add_job_type(self):
        """Add new job type"""
        job_name = self.new_job_entry.get().strip()
        
        if not job_name:
            messagebox.showwarning("คำเตือน", "กรุณาใส่ชื่อประเภทงาน")
            return
        
        try:
            self.db.execute_non_query("INSERT INTO job_types (job_name) VALUES (?)", (job_name,))
            self.new_job_entry.delete(0, tk.END)
            self.refresh_job_types()
            messagebox.showinfo("สำเร็จ", f"เพิ่มประเภทงาน '{job_name}' แล้ว")
        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถเพิ่มประเภทงานได้: {str(e)}")
    
    def delete_job_type(self):
        """Delete selected job type"""
        selection = self.job_listbox.curselection()
        if not selection:
            messagebox.showwarning("คำเตือน", "กรุณาเลือกประเภทงานที่จะลบ")
            return
        
        # Get the display text and extract job name
        display_text = self.job_listbox.get(selection[0])
        # Extract job name from "job_name (ID: X)" format
        job_name = display_text.split(" (ID: ")[0]
        job_id = self.job_types_data.get(job_name)
        
        if messagebox.askyesno("ยืนยัน", f"ต้องการลบประเภทงาน '{job_name}' หรือไม่?\n(จะลบ Dependencies ที่เกี่ยวข้องด้วย)"):
            try:
                # Delete dependencies first
                self.db.execute_non_query("DELETE FROM job_dependencies WHERE job_id = ? OR required_job_id = ?", 
                                         (job_id, job_id))
                # Delete job type
                self.db.execute_non_query("DELETE FROM job_types WHERE id = ?", (job_id,))
                self.refresh_job_types()
                # Reset selection
                self.current_selected_job = None
                self.selected_job_label.config(text="เลือกประเภทงานเพื่อตั้งค่า Dependencies")
                self.save_dependencies_btn.config(state=tk.DISABLED)
                messagebox.showinfo("สำเร็จ", f"ลบประเภทงาน '{job_name}' แล้ว")
            except Exception as e:
                messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถลบประเภทงานได้: {str(e)}")
    
    def process_barcode(self, event=None):
        """Process scanned barcode"""
        barcode = self.barcode_entry_var.get().strip()
        job_type = self.current_job_type.get()
        sub_job_type = self.current_sub_job_type.get()
        notes = self.notes_var.get().strip()
        
        if not barcode:
            return
        
        if not job_type:
            messagebox.showwarning("คำเตือน", "กรุณาเลือกประเภทงานหลักก่อน")
            return
        
        if not sub_job_type:
            messagebox.showwarning("คำเตือน", "กรุณาเลือกประเภทงานย่อยก่อน")
            return
        
        # Check for duplicates in same main job + sub job combination - no duplicates allowed
        try:
            # Get job IDs
            main_job_id = self.job_types_data.get(job_type)
            
            # Get sub job ID
            sub_job_query = """
                SELECT id FROM sub_job_types 
                WHERE main_job_id = ? AND sub_job_name = ? AND is_active = 1
            """
            sub_job_result = self.db.execute_query(sub_job_query, (main_job_id, sub_job_type))
            
            if not sub_job_result:
                messagebox.showerror("ข้อผิดพลาด", "ไม่พบประเภทงานย่อยที่เลือก")
                return
            
            sub_job_id = sub_job_result[0]['id']
            
            # Check for existing scan with same barcode + main job + sub job combination
            existing = self.db.execute_query(
                "SELECT scan_date, user_id FROM scan_logs WHERE barcode = ? AND job_id = ? AND sub_job_id = ? ORDER BY scan_date DESC", 
                (barcode, main_job_id, sub_job_id)
            )
            if existing:
                # Clear barcode immediately when duplicate is detected
                self.barcode_entry_var.set("")
                # Show duplicate warning with details - no option to continue
                self.show_duplicate_info(barcode, job_type, sub_job_type, existing[0])
                return
        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถตรวจสอบข้อมูลซ้ำได้: {str(e)}")
            return
        
        # Check dependencies validation AFTER duplicate check (using main job only)
        if not self.check_dependencies(barcode, job_type):
            return
        
        # Save barcode
        try:
            self.db.execute_non_query(
                "INSERT INTO scan_logs (barcode, scan_date, job_type, user_id, job_id, sub_job_id, notes) VALUES (?, GETDATE(), ?, ?, ?, ?, ?)",
                (barcode, job_type, self.db.current_user, main_job_id, sub_job_id, notes)
            )
            
            # Get the ID of the newly inserted record
            scan_id_result = self.db.execute_query("SELECT @@IDENTITY as scan_id")
            scan_id = scan_id_result[0]['scan_id'] if scan_id_result else "N/A"
            
            # Update status
            # current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            # status_text = f"สแกนสำเร็จ | ID: {scan_id} | บาร์โค้ด: {barcode} | งานหลัก: {job_type} | งานย่อย: {sub_job_type}"
            # if notes:
            #     status_text += f" | หมายเหตุ: {notes}"
            # status_text += f" | เวลา: {current_time}"
            
            # self.status_label.config(text=status_text, foreground="green")
            
            # Clear input for next scan
            self.barcode_entry_var.set("")
            # Keep job types and notes (user might want to scan multiple items with same settings)
            
            # Refresh scanning history table
            self.refresh_scanning_history()
            
            # Update today summary
            self.load_today_summary()
            
        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถบันทึกข้อมูลได้: {str(e)}")
            self.status_label.config(text="เกิดข้อผิดพลาดในการบันทึก", foreground="red")
    
    def check_dependencies(self, barcode, job_type):
        """Check if barcode can be scanned based on job dependencies"""
        try:
            current_job_id = self.job_types_data.get(job_type, 0)
            
            # Get required job IDs for current job type
            required_jobs_query = """
                SELECT jd.required_job_id, jt.job_name 
                FROM job_dependencies jd
                JOIN job_types jt ON jd.required_job_id = jt.id
                WHERE jd.job_id = ?
            """
            required_jobs = self.db.execute_query(required_jobs_query, (current_job_id,))
            
            if not required_jobs:
                # No dependencies, can scan
                return True
            
            # Check if all required jobs have been scanned for this barcode (excluding same job type)
            for required_job in required_jobs:
                required_job_id = required_job['required_job_id']
                required_job_name = required_job['job_name']
                
                # Check if this required job has been scanned for this barcode
                check_query = """
                    SELECT COUNT(*) as count 
                    FROM scan_logs 
                    WHERE barcode = ? AND job_id = ?
                """
                result = self.db.execute_query(check_query, (barcode, required_job_id))
                
                if result[0]['count'] == 0:
                    # Required job not found, show warning
                    self.status_label.config(
                        text=f"ไม่มีงาน {required_job_name}",
                        foreground="red"
                    )
                    messagebox.showwarning(
                        "คำเตือน",
                        f"ไม่มีงาน {required_job_name}"
                    )
                    self.barcode_entry_var.set("")  # Clear input
                    return False
            
            # All dependencies satisfied
            return True
            
        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", f"เกิดข้อผิดพลาดในการตรวจสอบ Dependencies: {str(e)}")
            return False
    
    def show_duplicate_info(self, barcode, job_type, sub_job_type, existing_record):
        """Show detailed duplicate barcode information dialog - no scanning allowed"""
        dialog = tk.Toplevel(self.root)
        dialog.title("ไม่อนุญาตให้สแกนซ้ำ")
        dialog.geometry("500x200")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.geometry("+%d+%d" % (self.root.winfo_rootx() + 50, self.root.winfo_rooty() + 50))
        
        # Main frame
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Format scan date
        scan_date = existing_record['scan_date']
        if scan_date:
            formatted_date = scan_date.strftime("%Y-%m-%d %H:%M:%S")
        else:
            formatted_date = "ไม่ทราบ"
        
        # Warning message with details
        warning_text = f"⚠️ ไม่อนุญาตให้สแกนซ้ำ ⚠️\n\n"
        warning_text += f"เลขพัสดุ: {barcode}\n"
        warning_text += f"ประเภทงานหลัก: {job_type}\n"
        warning_text += f"ประเภทงานย่อย: {sub_job_type}\n"
        warning_text += f"สแกนไปแล้วเมื่อ: {formatted_date}\n"
        warning_text += f"โดยผู้ใช้: {existing_record['user_id']}\n\n"
        warning_text += "กรุณาตรวจสอบเลขพัสดุอีกครั้ง"
        
        message_label = ttk.Label(main_frame, text=warning_text, 
                                 justify=tk.CENTER, font=("Arial", 11))
        message_label.pack(pady=20)
        
        # Button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=10)
        
        def close_dialog():
            dialog.destroy()
        
        # Close button
        close_btn = ttk.Button(button_frame, text="ปิด (Enter)", command=close_dialog)
        close_btn.pack()
        close_btn.focus_set()  # Set focus on button for Enter key
        
        # Bind Enter key to close dialog
        dialog.bind('<Return>', lambda e: close_dialog())
        dialog.bind('<Escape>', lambda e: close_dialog())
        
        # Auto-close after 3 seconds
        dialog.after(3000, close_dialog)
        
        # Set focus back to barcode entry after dialog
        def focus_barcode_entry():
            try:
                self.barcode_entry.focus_set()
            except:
                pass
        
        # Schedule focus return after dialog closes
        dialog.protocol("WM_DELETE_WINDOW", lambda: [close_dialog(), self.root.after(100, focus_barcode_entry)])
        
        # Wait for dialog to close
        self.root.wait_window(dialog)
        
        # Ensure barcode entry gets focus back
        self.root.after(100, focus_barcode_entry)
    
    def search_history(self, event=None):
        """Search and display history"""
        try:
            # Build query with filters
            query = """
                SELECT 
                    sl.id, 
                    sl.barcode, 
                    sl.scan_date, 
                    sl.job_type as main_job_name,
                    sjt.sub_job_name,
                    sl.notes,
                    sl.user_id 
                FROM scan_logs sl
                LEFT JOIN sub_job_types sjt ON sl.sub_job_id = sjt.id
                WHERE sl.scan_date >= DATEADD(day, -30, GETDATE())
            """
            params = []
            
            # Date filter
            date_str = self.date_entry.get().strip()
            if date_str:
                query += " AND CAST(sl.scan_date AS DATE) = ?"
                params.append(date_str)
            
            # Barcode filter
            barcode_search = self.search_barcode_entry.get().strip()
            if barcode_search:
                query += " AND sl.barcode LIKE ?"
                params.append(f"%{barcode_search}%")
            
            # Job type filter (main job only)
            job_search = self.search_job_combo.get().strip()
            if job_search:
                query += " AND sl.job_type = ?"
                params.append(job_search)
            
            query += " ORDER BY sl.scan_date DESC"
            
            # Execute query
            results = self.db.execute_query(query, tuple(params))
            
            # Clear existing data
            for item in self.history_tree.get_children():
                self.history_tree.delete(item)
            
            # Populate table
            for row in results:
                scan_date = row['scan_date'].strftime("%Y-%m-%d %H:%M:%S") if row['scan_date'] else ""
                sub_job_name = row['sub_job_name'] if row['sub_job_name'] else ""
                notes = row['notes'] if row['notes'] else ""
                
                self.history_tree.insert('', tk.END, values=(
                    row['id'],
                    row['barcode'],
                    scan_date,
                    row['main_job_name'],
                    sub_job_name,
                    notes,
                    row['user_id']
                ))
                
        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถค้นหาข้อมูลได้: {str(e)}")
    
    def refresh_report_job_types(self):
        """Refresh job types for report selection"""
        try:
            # Get job types
            query = "SELECT id, job_name FROM job_types ORDER BY job_name"
            results = self.db.execute_query(query)
            
            job_types = ["ทั้งหมด"] + [f"{row['id']} - {row['job_name']}" for row in results]
            self.report_job_type_combo['values'] = job_types
            
            # Store job types data for easier access
            self.report_job_types_data = {f"{row['id']} - {row['job_name']}": row['id'] for row in results}
            self.report_job_types_data["ทั้งหมด"] = None
            
        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถโหลดรายการงานหลักได้: {str(e)}")
    
    def on_report_job_type_change(self, event=None):
        """Handle job type change in report"""
        selected_job_type = self.report_job_type_var.get()
        job_type_id = self.report_job_types_data.get(selected_job_type)
        
        # Clear sub job type selection
        self.report_sub_job_type_var.set("")
        self.report_sub_job_type_combo['values'] = []
        
        if job_type_id:
            try:
                # Get sub job types for selected job type
                query = """
                    SELECT id, sub_job_name 
                    FROM sub_job_types 
                    WHERE main_job_id = ? AND is_active = 1 
                    ORDER BY sub_job_name
                """
                results = self.db.execute_query(query, (job_type_id,))
                
                sub_job_types = ["ทั้งหมด"] + [f"{row['id']} - {row['sub_job_name']}" for row in results]
                self.report_sub_job_type_combo['values'] = sub_job_types
                
                # Store sub job types data
                self.report_sub_job_types_data = {f"{row['id']} - {row['sub_job_name']}": row['id'] for row in results}
                self.report_sub_job_types_data["ทั้งหมด"] = None
                
            except Exception as e:
                messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถโหลดรายการงานรองได้: {str(e)}")
        else:
            # If "ทั้งหมด" is selected, load all sub job types
            try:
                query = """
                    SELECT id, sub_job_name 
                    FROM sub_job_types 
                    WHERE is_active = 1 
                    ORDER BY sub_job_name
                """
                results = self.db.execute_query(query)
                
                sub_job_types = ["ทั้งหมด"] + [f"{row['id']} - {row['sub_job_name']}" for row in results]
                self.report_sub_job_type_combo['values'] = sub_job_types
                
                self.report_sub_job_types_data = {f"{row['id']} - {row['sub_job_name']}": row['id'] for row in results}
                self.report_sub_job_types_data["ทั้งหมด"] = None
                
            except Exception as e:
                self.report_sub_job_types_data = {"ทั้งหมด": None}
    
    def run_report(self):
        """Generate report with filters like Web App"""
        # Get filter values
        report_date = self.report_date_var.get()
        selected_job_type = self.report_job_type_var.get()
        selected_sub_job_type = self.report_sub_job_type_var.get()
        note_filter = self.report_note_filter_var.get().strip()
        
        # Validate inputs
        if not report_date:
            messagebox.showwarning("คำเตือน", "กรุณาเลือกวันที่")
            return
        
        if not selected_job_type:
            messagebox.showwarning("คำเตือน", "กรุณาเลือกงานหลัก")
            return
        
        try:
            # Get actual IDs
            job_type_id = self.report_job_types_data.get(selected_job_type)
            sub_job_type_id = None
            if selected_sub_job_type and hasattr(self, 'report_sub_job_types_data'):
                sub_job_type_id = self.report_sub_job_types_data.get(selected_sub_job_type)
            
            # Build query similar to Web App
            if job_type_id and sub_job_type_id:
                # Specific job type and sub job type
                report_query = """
                    SELECT 
                        sl.barcode,
                        sl.scan_date,
                        sl.notes,
                        sl.user_id,
                        jt.job_name as job_type_name,
                        sjt.sub_job_name as sub_job_type_name
                    FROM scan_logs sl
                    LEFT JOIN job_types jt ON sl.job_id = jt.id
                    LEFT JOIN sub_job_types sjt ON sl.sub_job_id = sjt.id
                    WHERE sl.job_id = ? 
                    AND sl.sub_job_id = ?
                    AND CAST(sl.scan_date AS DATE) = ?
                """
                params = [job_type_id, sub_job_type_id, report_date]
            elif job_type_id:
                # Specific job type, all sub job types
                report_query = """
                    SELECT 
                        sl.barcode,
                        sl.scan_date,
                        sl.notes,
                        sl.user_id,
                        jt.job_name as job_type_name,
                        ISNULL(sjt.sub_job_name, 'ไม่มี') as sub_job_type_name
                    FROM scan_logs sl
                    LEFT JOIN job_types jt ON sl.job_id = jt.id
                    LEFT JOIN sub_job_types sjt ON sl.sub_job_id = sjt.id
                    WHERE sl.job_id = ? 
                    AND CAST(sl.scan_date AS DATE) = ?
                """
                params = [job_type_id, report_date]
            else:
                # All job types
                report_query = """
                    SELECT 
                        sl.barcode,
                        sl.scan_date,
                        sl.notes,
                        sl.user_id,
                        jt.job_name as job_type_name,
                        ISNULL(sjt.sub_job_name, 'ไม่มี') as sub_job_type_name
                    FROM scan_logs sl
                    LEFT JOIN job_types jt ON sl.job_id = jt.id
                    LEFT JOIN sub_job_types sjt ON sl.sub_job_id = sjt.id
                    WHERE CAST(sl.scan_date AS DATE) = ?
                """
                params = [report_date]
            
            # Add note filter if specified
            if note_filter:
                report_query += " AND sl.notes LIKE ?"
                params.append(f"%{note_filter}%")
            
            report_query += " ORDER BY sl.scan_date DESC"
            
            # Execute query
            results = self.db.execute_query(report_query, tuple(params))
            
            # Store data for export
            self.current_report_data = results
            self.current_report_summary = {
                'report_date': report_date,
                'job_type_name': selected_job_type,
                'sub_job_type_name': selected_sub_job_type or 'ทั้งหมด',
                'note_filter': note_filter if note_filter else None,
                'total_count': len(results),
                'generated_at': datetime.datetime.now().isoformat()
            }
            
            if not results:
                messagebox.showinfo("ผลลัพธ์", "ไม่พบข้อมูลในวันที่ที่เลือก")
                # Clear table
                for item in self.report_tree.get_children():
                    self.report_tree.delete(item)
                return
            
            # Clear existing data
            for item in self.report_tree.get_children():
                self.report_tree.delete(item)
            
            # Setup columns
            columns = ['barcode', 'scan_date', 'job_type_name', 'sub_job_type_name', 'user_id', 'notes']
            self.report_tree['columns'] = columns
            self.report_tree['show'] = 'headings'
            
            # Configure column headings and widths
            column_widths = {
                'barcode': 150,
                'scan_date': 150,
                'job_type_name': 120,
                'sub_job_type_name': 120,
                'user_id': 100,
                'notes': 200
            }
            
            column_names = {
                'barcode': 'บาร์โค้ด',
                'scan_date': 'วันที่/เวลา',
                'job_type_name': 'งานหลัก',
                'sub_job_type_name': 'งานรอง',
                'user_id': 'ผู้ใช้',
                'notes': 'หมายเหตุ'
            }
            
            for col in columns:
                self.report_tree.heading(col, text=column_names.get(col, col))
                self.report_tree.column(col, width=column_widths.get(col, 120))
            
            # Populate data
            for row in results:
                values = []
                for col in columns:
                    value = row.get(col, "")
                    if col == 'scan_date' and value:
                        # Format datetime
                        if isinstance(value, datetime.datetime):
                            value = value.strftime("%Y-%m-%d %H:%M:%S")
                    values.append(str(value) if value is not None else "")
                self.report_tree.insert('', tk.END, values=values)
            
            messagebox.showinfo("สำเร็จ", f"รันรายงานสำเร็จ พบข้อมูล {len(results)} รายการ")
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถรันรายงานได้: {str(e)}")
    
    def show_scan_context_menu(self, event):
        """Show context menu on right-click"""
        # Select the item under cursor
        item = self.scan_history_tree.identify_row(event.y)
        if item:
            self.scan_history_tree.selection_set(item)
            self.scan_context_menu.post(event.x_root, event.y_root)
    
    def edit_scan_record(self):
        """Edit selected scan record"""
        selection = self.scan_history_tree.selection()
        if not selection:
            messagebox.showwarning("คำเตือน", "กรุณาเลือกข้อมูลที่จะแก้ไข")
            return
        
        # Get selected record data
        item = selection[0]
        values = self.scan_history_tree.item(item, 'values')
        record_id = values[0]
        
        # Create edit dialog
        self.show_edit_dialog(record_id, values)
    
    def delete_scan_record(self):
        """Delete selected scan record"""
        selection = self.scan_history_tree.selection()
        if not selection:
            messagebox.showwarning("คำเตือน", "กรุณาเลือกข้อมูลที่จะลบ")
            return
        
        # Get selected record data
        item = selection[0]
        values = self.scan_history_tree.item(item, 'values')
        record_id = values[0]
        barcode = values[1]
        job_type = values[3]
        
        # Confirm deletion
        if messagebox.askyesno("ยืนยัน", 
                              f"ต้องการลบข้อมูลการสแกนหรือไม่?\n\nID: {record_id}\nบาร์โค้ด: {barcode}\nประเภทงาน: {job_type}"):
            try:
                # Delete from database
                self.db.execute_non_query("DELETE FROM scan_logs WHERE id = ?", (record_id,))
                messagebox.showinfo("สำเร็จ", "ลบข้อมูลเรียบร้อยแล้ว")
                
                # Refresh the history table
                self.refresh_scanning_history()
                
            except Exception as e:
                messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถลบข้อมูลได้: {str(e)}")
    
    def show_edit_dialog(self, record_id, current_values):
        """Show dialog for editing scan record"""
        dialog = tk.Toplevel(self.root)
        dialog.title("แก้ไขข้อมูลการสแกน")
        dialog.geometry("600x400")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.geometry("+%d+%d" % (self.root.winfo_rootx() + 100, self.root.winfo_rooty() + 100))
        
        # Main frame
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # ID (readonly)
        id_frame = ttk.Frame(main_frame)
        id_frame.pack(fill=tk.X, pady=5)
        ttk.Label(id_frame, text="ID:", width=15).pack(side=tk.LEFT)
        id_entry = ttk.Entry(id_frame, width=20, state="readonly")
        id_entry.pack(side=tk.LEFT, padx=10)
        id_entry.config(state="normal")
        id_entry.insert(0, current_values[0])
        id_entry.config(state="readonly")
        
        # Barcode
        barcode_frame = ttk.Frame(main_frame)
        barcode_frame.pack(fill=tk.X, pady=5)
        ttk.Label(barcode_frame, text="บาร์โค้ด:", width=15).pack(side=tk.LEFT)
        barcode_entry = ttk.Entry(barcode_frame, width=30)
        barcode_entry.pack(side=tk.LEFT, padx=10)
        barcode_entry.insert(0, current_values[1])
        
        # Scan date (readonly - informative)
        date_frame = ttk.Frame(main_frame)
        date_frame.pack(fill=tk.X, pady=5)
        ttk.Label(date_frame, text="วันที่/เวลา:", width=15).pack(side=tk.LEFT)
        date_entry = ttk.Entry(date_frame, width=30, state="readonly")
        date_entry.pack(side=tk.LEFT, padx=10)
        date_entry.config(state="normal")
        date_entry.insert(0, current_values[2])
        date_entry.config(state="readonly")
        
        # Main job type
        main_job_frame = ttk.Frame(main_frame)
        main_job_frame.pack(fill=tk.X, pady=5)
        ttk.Label(main_job_frame, text="ประเภทงานหลัก:", width=15).pack(side=tk.LEFT)
        main_job_var = tk.StringVar()
        main_job_combo = ttk.Combobox(main_job_frame, textvariable=main_job_var, state="readonly", width=25)
        main_job_combo.pack(side=tk.LEFT, padx=10)
        
        # Sub job type
        sub_job_frame = ttk.Frame(main_frame)
        sub_job_frame.pack(fill=tk.X, pady=5)
        ttk.Label(sub_job_frame, text="ประเภทงานย่อย:", width=15).pack(side=tk.LEFT)
        sub_job_var = tk.StringVar()
        sub_job_combo = ttk.Combobox(sub_job_frame, textvariable=sub_job_var, state="readonly", width=25)
        sub_job_combo.pack(side=tk.LEFT, padx=10)
        
        # Notes
        notes_frame = ttk.Frame(main_frame)
        notes_frame.pack(fill=tk.X, pady=5)
        ttk.Label(notes_frame, text="หมายเหตุ:", width=15).pack(side=tk.LEFT)
        notes_entry = ttk.Entry(notes_frame, width=40)
        notes_entry.pack(side=tk.LEFT, padx=10)
        if len(current_values) > 5:
            notes_entry.insert(0, current_values[5] or "")
        
        # User (readonly - informative)
        user_frame = ttk.Frame(main_frame)
        user_frame.pack(fill=tk.X, pady=5)
        ttk.Label(user_frame, text="ผู้ใช้:", width=15).pack(side=tk.LEFT)
        user_entry = ttk.Entry(user_frame, width=30, state="readonly")
        user_entry.pack(side=tk.LEFT, padx=10)
        user_entry.config(state="normal")
        user_entry.insert(0, current_values[6] if len(current_values) > 6 else current_values[4])
        user_entry.config(state="readonly")
        
        # Load job types for combo
        try:
            job_types = list(self.job_types_data.keys())
            main_job_combo['values'] = job_types
            main_job_var.set(current_values[3])  # Main job name
            
            # Bind main job change to update sub jobs
            def on_main_job_change_edit(event=None):
                selected_main_job = main_job_var.get()
                main_job_id = self.job_types_data.get(selected_main_job)
                
                if main_job_id:
                    try:
                        query = """
                            SELECT sub_job_name 
                            FROM sub_job_types 
                            WHERE main_job_id = ? AND is_active = 1
                            ORDER BY sub_job_name
                        """
                        results = self.db.execute_query(query, (main_job_id,))
                        sub_job_names = [row['sub_job_name'] for row in results]
                        sub_job_combo['values'] = sub_job_names
                        
                        # Clear current sub job selection if main job changed
                        if event:  # Only clear if this was triggered by user change
                            sub_job_var.set('')
                    except Exception as e:
                        print(f"Error loading sub jobs in edit dialog: {str(e)}")
                        sub_job_combo['values'] = []
            
            main_job_combo.bind('<<ComboboxSelected>>', on_main_job_change_edit)
            
            # Load initial sub jobs and set current sub job
            on_main_job_change_edit()  # Load sub jobs for current main job
            if len(current_values) > 4:
                sub_job_var.set(current_values[4] or "")  # Sub job name
            
        except Exception as e:
            print(f"Error setting up edit dialog: {str(e)}")
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=20)
        
        def save_changes():
            new_barcode = barcode_entry.get().strip()
            new_main_job = main_job_var.get()
            new_sub_job = sub_job_var.get()
            new_notes = notes_entry.get().strip()
            
            if not new_barcode:
                messagebox.showwarning("คำเตือน", "กรุณาใส่บาร์โค้ด")
                return
            
            if not new_main_job:
                messagebox.showwarning("คำเตือน", "กรุณาเลือกประเภทงานหลัก")
                return
            
            if not new_sub_job:
                messagebox.showwarning("คำเตือน", "กรุณาเลือกประเภทงานย่อย")
                return
            
            try:
                # Get job IDs
                new_main_job_id = self.job_types_data.get(new_main_job, 0)
                
                # Get sub job ID
                sub_job_query = """
                    SELECT id FROM sub_job_types 
                    WHERE main_job_id = ? AND sub_job_name = ? AND is_active = 1
                """
                sub_job_result = self.db.execute_query(sub_job_query, (new_main_job_id, new_sub_job))
                
                if not sub_job_result:
                    messagebox.showerror("ข้อผิดพลาด", "ไม่พบประเภทงานย่อยที่เลือก")
                    return
                
                new_sub_job_id = sub_job_result[0]['id']
                
                # Update database
                self.db.execute_non_query(
                    "UPDATE scan_logs SET barcode = ?, job_type = ?, job_id = ?, sub_job_id = ?, notes = ? WHERE id = ?",
                    (new_barcode, new_main_job, new_main_job_id, new_sub_job_id, new_notes, record_id)
                )
                
                messagebox.showinfo("สำเร็จ", "อัพเดทข้อมูลเรียบร้อยแล้ว")
                dialog.destroy()
                
                # Refresh the history table
                self.refresh_scanning_history()
                
            except Exception as e:
                messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถอัพเดทข้อมูลได้: {str(e)}")
        
        def cancel_edit():
            dialog.destroy()
        
        ttk.Button(button_frame, text="บันทึกการเปลี่ยนแปลง", command=save_changes).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="ยกเลิก", command=cancel_edit).pack(side=tk.LEFT, padx=5)
        
        # Focus on barcode entry
        barcode_entry.focus_set()
        barcode_entry.select_range(0, tk.END)
    
    def export_report(self):
        """Export current report data to Excel with summary"""
        if not self.current_report_data:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลสำหรับส่งออก กรุณารันรายงานก่อน")
            return
        
        try:
            # Create filename with current date
            today = datetime.date.today().strftime("%Y%m%d")
            default_filename = f"รายงาน_{today}.xlsx"
            
            # Ask user for save location
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=default_filename
            )
            
            if filename:
                # Create Excel workbook with multiple sheets
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    
                    # Summary sheet
                    if hasattr(self, 'current_report_summary'):
                        summary_data = {
                            'รายการ': [
                                'วันที่รายงาน',
                                'งานหลัก', 
                                'งานรอง',
                                'กรองหมายเหตุ',
                                'จำนวนรวม',
                                'วันที่สร้างรายงาน'
                            ],
                            'ค่า': [
                                self.current_report_summary.get('report_date', ''),
                                self.current_report_summary.get('job_type_name', ''),
                                self.current_report_summary.get('sub_job_type_name', ''),
                                self.current_report_summary.get('note_filter', 'ไม่มี') or 'ไม่มี',
                                self.current_report_summary.get('total_count', 0),
                                datetime.datetime.fromisoformat(self.current_report_summary.get('generated_at', '')).strftime('%Y-%m-%d %H:%M:%S') if self.current_report_summary.get('generated_at') else ''
                            ]
                        }
                        summary_df = pd.DataFrame(summary_data)
                        summary_df.to_excel(writer, sheet_name='สรุป', index=False)
                    
                    # Detail data sheet
                    # Convert data and rename columns to Thai
                    df = pd.DataFrame(self.current_report_data)
                    if not df.empty:
                        # Rename columns to Thai
                        column_mapping = {
                            'barcode': 'บาร์โค้ด',
                            'scan_date': 'วันที่/เวลา',
                            'job_type_name': 'งานหลัก',
                            'sub_job_type_name': 'งานรอง',
                            'user_id': 'ผู้ใช้',
                            'notes': 'หมายเหตุ'
                        }
                        
                        # Rename only existing columns
                        df_columns = {col: column_mapping.get(col, col) for col in df.columns if col in column_mapping}
                        df = df.rename(columns=df_columns)
                        
                        # Format datetime columns
                        for col in df.columns:
                            if df[col].dtype == 'datetime64[ns]' or 'date' in col.lower():
                                df[col] = pd.to_datetime(df[col], errors='ignore').dt.strftime('%Y-%m-%d %H:%M:%S')
                        
                        df.to_excel(writer, sheet_name='รายละเอียด', index=False)
                    
                messagebox.showinfo("สำเร็จ", f"ส่งออกไฟล์สำเร็จ\n{filename}\n\nประกอบด้วย:\n- แผ่นสรุป: ข้อมูลสรุปรายงาน\n- แผ่นรายละเอียด: ข้อมูลทั้งหมด")
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถส่งออกไฟล์ได้: {str(e)}")


    
    # Import/Export functionality
    def download_template(self):
        """Download Excel template for data import"""
        try:
            # Get job types data for template
            job_types_data = []
            sub_job_types_data = []
            
            try:
                # Get job types with IDs
                job_results = self.db.execute_query("SELECT id, job_name FROM job_types ORDER BY job_name")
                for job in job_results:
                    job_types_data.append(f"{job['id']} - {job['job_name']}")
                
                # Get sub job types with IDs
                sub_job_results = self.db.execute_query("""
                    SELECT sjt.id, sjt.sub_job_name, jt.job_name as main_job_name, sjt.main_job_id
                    FROM sub_job_types sjt
                    JOIN job_types jt ON sjt.main_job_id = jt.id
                    WHERE sjt.is_active = 1
                    ORDER BY jt.job_name, sjt.sub_job_name
                """)
                for sub_job in sub_job_results:
                    sub_job_types_data.append(f"{sub_job['id']} - {sub_job['sub_job_name']} (งานหลัก: {sub_job['main_job_name']})")
                    
            except Exception as e:
                print(f"Error loading job types: {str(e)}")
            
            # Create sample data for template
            template_data = {
                'บาร์โค้ด': ['SAMPLE001', 'SAMPLE002', 'SAMPLE003'],
                'ID_ประเภทงานหลัก': ['', '', ''],
                'ID_ประเภทงานย่อย': ['', '', ''],
                'หมายเหตุ': ['ตัวอย่างหมายเหตุ 1', '', 'ตัวอย่างหมายเหตุ 3']
            }
            
            # Create DataFrame
            df = pd.DataFrame(template_data)
            
            # Ask for save location
            file_path = filedialog.asksaveasfilename(
                title="บันทึกเทมเพลต Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"WMS_Import_Template_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
            )
            
            if file_path:
                # Save to Excel with formatting
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    # Write main data sheet
                    df.to_excel(writer, sheet_name='ข้อมูลนำเข้า', index=False)
                    
                    # Get the workbook and worksheet objects
                    workbook = writer.book
                    worksheet = writer.sheets['ข้อมูลนำเข้า']
                    
                    # Import openpyxl styling modules
                    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter
                    
                    # Define styles
                    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    blue_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                    header_font = Font(bold=True, size=12)
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Format headers
                    for col_num, column in enumerate(df.columns, 1):
                        cell = worksheet.cell(row=1, column=col_num)
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                        
                        # Color required columns red, optional blue
                        if column in ['บาร์โค้ด', 'ID_ประเภทงานหลัก', 'ID_ประเภทงานย่อย']:
                            cell.fill = red_fill
                        else:
                            cell.fill = blue_fill
                    
                    # Set column widths
                    column_widths = {
                        'A': 20,  # บาร์โค้ด
                        'B': 25,  # ID_ประเภทงานหลัก
                        'C': 25,  # ID_ประเภทงานย่อย
                        'D': 30   # หมายเหตุ
                    }
                    
                    for col_letter, width in column_widths.items():
                        worksheet.column_dimensions[col_letter].width = width
                    
                    # Format data rows
                    for row in range(2, len(df) + 2):
                        for col in range(1, len(df.columns) + 1):
                            cell = worksheet.cell(row=row, column=col)
                            cell.border = border
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Add instruction sheet
                    instructions_data = {
                        'คำแนะนำการใช้งาน': [
                            '1. ใส่ข้อมูลในแท็บ "ข้อมูลนำเข้า"',
                            '2. คอลัมน์สีแดง = บังคับกรอก, คอลัมน์สีฟ้า = ไม่บังคับ',
                            '3. ใช้ ID แทนชื่อสำหรับประเภทงาน (ดูรายการด้านล่าง)',
                            '4. บันทึกไฟล์และนำเข้าผ่านระบบ',
                            '',
                            'หมายเหตุ:',
                            '- ระบบจะตรวจสอบข้อมูลซ้ำกัน',
                            '- ระบบจะตรวจสอบ Dependencies',
                            '- ไฟล์ต้องเป็น .xlsx เท่านั้น',
                            '- ID_ประเภทงานย่อย ต้องสัมพันธ์กับ ID_ประเภทงานหลัก'
                        ]
                    }
                    
                    instructions_df = pd.DataFrame(instructions_data)
                    instructions_df.to_excel(writer, sheet_name='คำแนะนำ', index=False)
                    
                    # Format instructions sheet
                    inst_worksheet = writer.sheets['คำแนะนำ']
                    inst_worksheet.column_dimensions['A'].width = 60
                    
                    for row in range(1, len(instructions_df) + 2):
                        cell = inst_worksheet.cell(row=row, column=1)
                        if row == 1:
                            cell.font = header_font
                            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        cell.border = border
                    
                    # Add job types reference sheet
                    if job_types_data:
                        job_ref_data = {
                            'รายการประเภทงานหลัก (ใช้ ID)': job_types_data
                        }
                        
                        job_ref_df = pd.DataFrame(job_ref_data)
                        job_ref_df.to_excel(writer, sheet_name='รายการประเภทงานหลัก', index=False)
                        
                        # Format job types sheet
                        job_worksheet = writer.sheets['รายการประเภทงานหลัก']
                        job_worksheet.column_dimensions['A'].width = 50
                        
                        # Format header
                        cell = job_worksheet.cell(row=1, column=1)
                        cell.font = header_font
                        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                        
                        # Format data
                        for row in range(2, len(job_ref_df) + 2):
                            cell = job_worksheet.cell(row=row, column=1)
                            cell.border = border
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Add sub job types reference sheet
                    if sub_job_types_data:
                        sub_job_ref_data = {
                            'รายการประเภทงานย่อย (ใช้ ID)': sub_job_types_data
                        }
                        
                        sub_job_ref_df = pd.DataFrame(sub_job_ref_data)
                        sub_job_ref_df.to_excel(writer, sheet_name='รายการประเภทงานย่อย', index=False)
                        
                        # Format sub job types sheet
                        sub_job_worksheet = writer.sheets['รายการประเภทงานย่อย']
                        sub_job_worksheet.column_dimensions['A'].width = 60
                        
                        # Format header
                        cell = sub_job_worksheet.cell(row=1, column=1)
                        cell.font = header_font
                        cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                        
                        # Format data
                        for row in range(2, len(sub_job_ref_df) + 2):
                            cell = sub_job_worksheet.cell(row=row, column=1)
                            cell.border = border
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                
                messagebox.showinfo("สำเร็จ", f"ดาวน์โหลดเทมเพลตเรียบร้อยแล้ว\n{file_path}\n\nคำแนะนำ:\n- คอลัมน์สีแดง = บังคับกรอก\n- คอลัมน์สีฟ้า = ไม่บังคับ\n- ใช้ ID จากแท็บรายการประเภทงาน")
                
        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถสร้างเทมเพลตได้: {str(e)}")
    
    def show_sample_data(self):
        """Show sample data format dialog"""
        dialog = tk.Toplevel(self.root)
        dialog.title("ตัวอย่างรูปแบบข้อมูล")
        dialog.geometry("700x500")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.geometry("+%d+%d" % (self.root.winfo_rootx() + 100, self.root.winfo_rooty() + 50))
        
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Sample data info
        info_text = """ตัวอย่างรูปแบบข้อมูลที่ถูกต้อง:

คอลัมน์ที่ต้องมี:
✓ บาร์โค้ด - รหัสบาร์โค้ดของสินค้า (บังคับ) [สีแดง]
✓ ID_ประเภทงานหลัก - ใช้ ID เท่านั้น เช่น 1, 2, 3 (บังคับ) [สีแดง]
✓ ID_ประเภทงานย่อย - ใช้ ID เท่านั้น เช่น 10, 11, 12 (บังคับ) [สีแดง]
○ หมายเหตุ - ข้อมูลเพิ่มเติม (ไม่บังคับ) [สีฟ้า]

ตัวอย่างข้อมูล:
บาร์โค้ด: PROD001
ID_ประเภทงานหลัก: 1
ID_ประเภทงานย่อย: 10
หมายเหตุ: ตัวอย่างหมายเหตุ"""
        
        # Add current job types with IDs
        try:
            info_text += "\n\n=== รายการประเภทงานหลัก ===\n"
            job_results = self.db.execute_query("SELECT id, job_name FROM job_types ORDER BY id")
            for job in job_results:
                info_text += f"ID {job['id']}: {job['job_name']}\n"
            
            info_text += "\n=== รายการประเภทงานย่อย ===\n"
            sub_job_results = self.db.execute_query("""
                SELECT sjt.id, sjt.sub_job_name, jt.job_name as main_job_name
                FROM sub_job_types sjt
                JOIN job_types jt ON sjt.main_job_id = jt.id
                WHERE sjt.is_active = 1
                ORDER BY sjt.id
            """)
            for sub_job in sub_job_results:
                info_text += f"ID {sub_job['id']}: {sub_job['sub_job_name']} (งานหลัก: {sub_job['main_job_name']})\n"
                
        except Exception as e:
            info_text += f"\n\nไม่สามารถโหลดรายการประเภทงานได้: {str(e)}"
        
        # Create scrollable text widget
        text_frame = ttk.Frame(main_frame)
        text_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        text_widget = tk.Text(text_frame, wrap=tk.WORD, font=("Arial", 10))
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        text_widget.insert('1.0', info_text)
        text_widget.config(state=tk.DISABLED)
        
        # Close button
        ttk.Button(main_frame, text="ปิด", command=dialog.destroy).pack(pady=20)
    
    def select_import_file(self):
        """Select Excel file for import"""
        file_path = filedialog.askopenfilename(
            title="เลือกไฟล์ Excel สำหรับนำเข้า",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if file_path:
            self.import_file_var.set(file_path)
            self.import_status_label.config(text="เลือกไฟล์แล้ว - กดตรวจสอบข้อมูลเพื่อดูตัวอย่าง", 
                                          foreground="blue")
            # Clear previous data
            self.import_data_df = None
            self.import_validation_results = None
            self.clear_import_preview()
    
    def clear_import_data(self):
        """Clear import data and reset"""
        self.import_file_var.set("")
        self.import_data_df = None
        self.import_validation_results = None
        self.clear_import_preview()
        self.import_status_label.config(text="เลือกไฟล์ Excel เพื่อเริ่มต้น", foreground="blue")
    
    def clear_import_preview(self):
        """Clear import preview table"""
        for item in self.import_preview_tree.get_children():
            self.import_preview_tree.delete(item)
    
    def validate_import_data(self):
        """Validate import data from Excel file"""
        file_path = self.import_file_var.get()
        if not file_path:
            messagebox.showwarning("คำเตือน", "กรุณาเลือกไฟล์ Excel ก่อน")
            return
        
        try:
            # Read Excel file
            df = pd.read_excel(file_path, sheet_name=0)
            
            # Check required columns
            required_columns = ['บาร์โค้ด', 'ID_ประเภทงานหลัก', 'ID_ประเภทงานย่อย']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                messagebox.showerror("ข้อผิดพลาด", 
                                   f"ไฟล์ Excel ไม่มีคอลัมน์ที่จำเป็น:\n{', '.join(missing_columns)}\n\n" +
                                   "คอลัมน์ที่ต้องมี:\n- บาร์โค้ด\n- ID_ประเภทงานหลัก\n- ID_ประเภทงานย่อย\n- หมายเหตุ (ไม่บังคับ)")
                return
            
            # Clean and validate data
            self.import_data_df = df.copy()
            self.import_validation_results = []
            
            self.clear_import_preview()
            
            for idx, row in df.iterrows():
                validation_result = self.validate_import_row(row, idx + 1)
                self.import_validation_results.append(validation_result)
                
                # Add to preview table
                status = "✓" if validation_result['valid'] else "✗"
                error_msg = validation_result.get('errors', '')
                
                # Get job names for display
                main_job_name = validation_result.get('main_job_name', row.get('ID_ประเภทงานหลัก', ''))
                sub_job_name = validation_result.get('sub_job_name', row.get('ID_ประเภทงานย่อย', ''))
                
                # Configure row color based on status
                item = self.import_preview_tree.insert('', tk.END, values=(
                    status,
                    row.get('บาร์โค้ด', ''),
                    f"{row.get('ID_ประเภทงานหลัก', '')} - {main_job_name}",
                    f"{row.get('ID_ประเภทงานย่อย', '')} - {sub_job_name}",
                    row.get('หมายเหตุ', ''),
                    error_msg
                ))
                
                # Color coding
                if not validation_result['valid']:
                    self.import_preview_tree.set(item, 'สถานะ', '✗')
                else:
                    self.import_preview_tree.set(item, 'สถานะ', '✓')
            
            # Update status
            valid_count = sum(1 for r in self.import_validation_results if r['valid'])
            total_count = len(self.import_validation_results)
            
            if valid_count == total_count:
                self.import_status_label.config(
                    text=f"ตรวจสอบเรียบร้อย: {valid_count}/{total_count} แถว พร้อมนำเข้า", 
                    foreground="green"
                )
            else:
                self.import_status_label.config(
                    text=f"พบข้อผิดพลาด: {valid_count}/{total_count} แถวที่ถูกต้อง", 
                    foreground="red"
                )
                
        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", f"ไม่สามารถอ่านไฟล์ Excel ได้: {str(e)}")
    
    def validate_import_row(self, row, row_number):
        """Validate a single row from import data"""
        result = {'valid': True, 'errors': [], 'row_number': row_number}
        
        # Check required fields
        barcode = str(row.get('บาร์โค้ด', '')).strip()
        main_job_id_raw = str(row.get('ID_ประเภทงานหลัก', '')).strip()
        sub_job_id_raw = str(row.get('ID_ประเภทงานย่อย', '')).strip()
        
        if not barcode or barcode == 'nan':
            result['valid'] = False
            result['errors'].append('ไม่มีบาร์โค้ด')
        
        if not main_job_id_raw or main_job_id_raw == 'nan':
            result['valid'] = False
            result['errors'].append('ไม่มี ID ประเภทงานหลัก')
        
        if not sub_job_id_raw or sub_job_id_raw == 'nan':
            result['valid'] = False
            result['errors'].append('ไม่มี ID ประเภทงานย่อย')
        
        # Validate job IDs are numeric
        main_job_id = None
        sub_job_id = None
        
        try:
            if main_job_id_raw and main_job_id_raw != 'nan':
                main_job_id = int(float(main_job_id_raw))
        except (ValueError, TypeError):
            result['valid'] = False
            result['errors'].append(f'ID ประเภทงานหลักไม่ถูกต้อง: {main_job_id_raw}')
        
        try:
            if sub_job_id_raw and sub_job_id_raw != 'nan':
                sub_job_id = int(float(sub_job_id_raw))
        except (ValueError, TypeError):
            result['valid'] = False
            result['errors'].append(f'ID ประเภทงานย่อยไม่ถูกต้อง: {sub_job_id_raw}')
        
        # Validate job types exist in database
        main_job_name = ''
        sub_job_name = ''
        
        if main_job_id is not None:
            try:
                job_check = self.db.execute_query("SELECT job_name FROM job_types WHERE id = ?", (main_job_id,))
                if not job_check:
                    result['valid'] = False
                    result['errors'].append(f'ไม่พบประเภทงานหลัก ID: {main_job_id}')
                else:
                    main_job_name = job_check[0]['job_name']
                    result['main_job_name'] = main_job_name
            except Exception as e:
                result['valid'] = False
                result['errors'].append(f'ไม่สามารถตรวจสอบประเภทงานหลักได้: {str(e)}')
        
        if sub_job_id is not None:
            try:
                sub_job_check = self.db.execute_query("""
                    SELECT sjt.sub_job_name, sjt.main_job_id, jt.job_name as main_job_name
                    FROM sub_job_types sjt
                    JOIN job_types jt ON sjt.main_job_id = jt.id
                    WHERE sjt.id = ? AND sjt.is_active = 1
                """, (sub_job_id,))
                
                if not sub_job_check:
                    result['valid'] = False
                    result['errors'].append(f'ไม่พบประเภทงานย่อย ID: {sub_job_id}')
                else:
                    sub_job_name = sub_job_check[0]['sub_job_name']
                    sub_job_main_id = sub_job_check[0]['main_job_id']
                    result['sub_job_name'] = sub_job_name
                    
                    # Check if sub job belongs to main job
                    if main_job_id is not None and sub_job_main_id != main_job_id:
                        result['valid'] = False
                        result['errors'].append(f'ประเภทงานย่อย ID {sub_job_id} ไม่สัมพันธ์กับประเภทงานหลัก ID {main_job_id}')
                        
            except Exception as e:
                result['valid'] = False
                result['errors'].append(f'ไม่สามารถตรวจสอบประเภทงานย่อยได้: {str(e)}')
        
        # Check for duplicates if option enabled
        if self.skip_duplicates_var.get() and barcode and main_job_id and sub_job_id:
            try:
                # Check for existing record
                existing = self.db.execute_query(
                    "SELECT COUNT(*) as count FROM scan_logs WHERE barcode = ? AND job_id = ? AND sub_job_id = ?", 
                    (barcode, main_job_id, sub_job_id)
                )
                
                if existing and existing[0]['count'] > 0:
                    result['valid'] = False
                    result['errors'].append('ข้อมูลซ้ำในระบบ')
                    
            except Exception as e:
                result['errors'].append(f'ไม่สามารถตรวจสอบข้อมูลซ้ำได้: {str(e)}')
        
        # Check dependencies if option enabled
        if self.validate_dependencies_var.get() and main_job_name and result['valid']:
            try:
                if not self.check_dependencies(barcode, main_job_name):
                    result['valid'] = False
                    result['errors'].append('ไม่ผ่านการตรวจสอบ Dependencies')
            except:
                pass
        
        result['errors'] = '; '.join(result['errors'])
        return result
    
    def import_data(self):
        """Import validated data into database"""
        if not self.import_validation_results:
            messagebox.showwarning("คำเตือน", "กรุณาตรวจสอบข้อมูลก่อนนำเข้า")
            return
        
        # Count valid records
        valid_records = [r for r in self.import_validation_results if r['valid']]
        
        if not valid_records:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลที่ถูกต้องสำหรับการนำเข้า")
            return
        
        # Confirm import
        total_count = len(self.import_validation_results)
        valid_count = len(valid_records)
        
        if valid_count < total_count:
            if not messagebox.askyesno("ยืนยัน", 
                                     f"พบข้อมูลที่ถูกต้อง {valid_count} จาก {total_count} แถว\n" +
                                     "ต้องการนำเข้าเฉพาะข้อมูลที่ถูกต้องหรือไม่?"):
                return
        else:
            if not messagebox.askyesno("ยืนยัน", 
                                     f"ต้องการนำเข้าข้อมูล {valid_count} แถวหรือไม่?"):
                return
        
        try:
            success_count = 0
            error_count = 0
            
            for validation_result in self.import_validation_results:
                if not validation_result['valid']:
                    continue
                
                try:
                    row_idx = validation_result['row_number'] - 1
                    row = self.import_data_df.iloc[row_idx]
                    
                    barcode = str(row['บาร์โค้ด']).strip()
                    main_job_id = int(float(str(row['ID_ประเภทงานหลัก']).strip()))
                    sub_job_id = int(float(str(row['ID_ประเภทงานย่อย']).strip()))
                    notes = str(row.get('หมายเหตุ', '')).strip()
                    
                    # Get job name for job_type field (for compatibility)
                    job_name_result = self.db.execute_query("SELECT job_name FROM job_types WHERE id = ?", (main_job_id,))
                    main_job_name = job_name_result[0]['job_name'] if job_name_result else f"ID_{main_job_id}"
                    
                    # Insert into database
                    self.db.execute_non_query(
                        "INSERT INTO scan_logs (barcode, scan_date, job_type, user_id, job_id, sub_job_id, notes) VALUES (?, GETDATE(), ?, ?, ?, ?, ?)",
                        (barcode, main_job_name, f"{self.db.current_user}_IMPORT", main_job_id, sub_job_id, notes)
                    )
                    
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    print(f"Error importing row {validation_result['row_number']}: {str(e)}")
            
            # Show results
            if error_count == 0:
                messagebox.showinfo("สำเร็จ", f"นำเข้าข้อมูลเรียบร้อย {success_count} แถว")
            else:
                messagebox.showwarning("เสร็จสิ้น", 
                                     f"นำเข้าข้อมูลเสร็จสิ้น\n" +
                                     f"สำเร็จ: {success_count} แถว\n" +
                                     f"ข้อผิดพลาด: {error_count} แถว")
            
            # Refresh scanning history if on scanning tab
            try:
                self.refresh_scanning_history()
            except:
                pass
            
            # Update status
            self.import_status_label.config(
                text=f"นำเข้าเรียบร้อย: {success_count} แถว", 
                foreground="green"
            )
            
        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", f"เกิดข้อผิดพลาดในการนำเข้าข้อมูล: {str(e)}")
    
    def load_today_summary(self):
        """Load today's scan summary data"""
        try:
            # Check if UI is ready
            if not hasattr(self, 'summary_labels') or not self.summary_labels:
                return
            job_type_name = self.current_job_type.get()
            sub_job_type_name = self.current_sub_job_type.get()
            notes_filter = self.notes_var.get().strip()
            
            # If no job type selected, clear summary
            if not job_type_name:
                self.clear_summary()
                return
            
            # Get job type ID
            job_result = self.db.execute_query("SELECT id FROM job_types WHERE job_name = ?", (job_type_name,))
            if not job_result:
                self.clear_summary()
                return
            
            job_type_id = job_result[0]['id']
            sub_job_id = None
            
            # Get sub job type ID if selected
            if sub_job_type_name:
                sub_result = self.db.execute_query(
                    "SELECT id FROM sub_job_types WHERE sub_job_name = ? AND main_job_id = ?", 
                    (sub_job_type_name, job_type_id)
                )
                if sub_result:
                    sub_job_id = sub_result[0]['id']
            
            # Build query for today's count
            if sub_job_id:
                # With sub job type
                count_query = """
                    SELECT COUNT(*) as total_count
                    FROM scan_logs
                    WHERE job_id = ? 
                    AND sub_job_id = ?
                    AND CAST(scan_date AS DATE) = CAST(GETDATE() AS DATE)
                """
                params = [job_type_id, sub_job_id]
            else:
                # Without sub job type - count all for this job type
                count_query = """
                    SELECT COUNT(*) as total_count
                    FROM scan_logs
                    WHERE job_id = ? 
                    AND CAST(scan_date AS DATE) = CAST(GETDATE() AS DATE)
                """
                params = [job_type_id]
            
            # Add notes filter if specified
            if notes_filter:
                count_query += " AND notes LIKE ?"
                params.append(f"%{notes_filter}%")
            
            result = self.db.execute_query(count_query, tuple(params))
            total_count = result[0]['total_count'] if result else 0
            
            # Update summary display
            self.update_summary_display(job_type_name, sub_job_type_name or "ไม่มี", 
                                      notes_filter or None, total_count)
            
        except Exception as e:
            print(f"Error loading today summary: {str(e)}")
            self.clear_summary()
    
    def update_summary_display(self, job_type, sub_job_type, notes_filter, count):
        """Update the summary display with new data"""
        try:
            # Update job type
            self.summary_labels['job_type'].config(text=job_type)
            
            # Update sub job type
            self.summary_labels['sub_job_type'].config(text=sub_job_type)
            
            # Update notes filter
            if notes_filter:
                self.summary_labels['notes_filter'].config(text=f'"{notes_filter}"')
            else:
                self.summary_labels['notes_filter'].config(text="ไม่มี")
            
            # Update count with prominent display
            self.summary_labels['count'].config(text=str(count))
            
            # Update last update time
            import datetime
            now = datetime.datetime.now()
            self.summary_labels['last_update'].config(text=now.strftime("%d/%m/%Y %H:%M:%S"))
            
            # Show the summary frame
            self.today_summary_frame.pack(fill=tk.X, pady=10)
            
        except Exception as e:
            print(f"Error updating summary display: {str(e)}")
    
    def clear_summary(self):
        """Clear the summary display"""
        try:
            self.summary_labels['job_type'].config(text="ไม่ได้เลือก")
            self.summary_labels['sub_job_type'].config(text="ไม่ได้เลือก")
            self.summary_labels['notes_filter'].config(text="ไม่มี")
            self.summary_labels['count'].config(text="0")
            self.summary_labels['last_update'].config(text="ยังไม่มีข้อมูล")
            
            # Hide the summary frame
            self.today_summary_frame.pack_forget()
            
        except Exception as e:
            print(f"Error clearing summary: {str(e)}")
    
    def toggle_fullscreen(self, event=None):
        """สลับโหมดเต็มจอ"""
        try:
            current_state = self.root.attributes('-fullscreen')
            self.root.attributes('-fullscreen', not current_state)
            
            if not current_state:
                # เข้าสู่โหมดเต็มจอ - ซ่อน title bar
                self.root.overrideredirect(True)
            else:
                # ออกจากโหมดเต็มจอ - แสดง title bar
                self.root.overrideredirect(False)
                
        except Exception as e:
            # ถ้าไม่รองรับ fullscreen ให้ใช้ zoomed แทน
            try:
                if self.root.state() == 'zoomed':
                    self.root.state('normal')
                else:
                    self.root.state('zoomed')
            except Exception:
                print(f"ไม่สามารถเปลี่ยนโหมดเต็มจอได้: {e}")
    
    def toggle_maximize(self, event=None):
        """สลับโหมด maximize"""
        try:
            if self.root.state() == 'zoomed':
                self.root.state('normal')
            else:
                self.root.state('zoomed')
        except Exception as e:
            print(f"ไม่สามารถ maximize/restore ได้: {e}")

def main():
    """Main application entry point"""
    # Show login window first
    login = LoginWindow()
    connection_info = login.run()
    
    if not connection_info:
        # User cancelled login
        print("Login cancelled. Exiting...")
        return
    
    # Create main application window
    root = tk.Tk()
    app = WMSScannerApp(root, connection_info)
    
    # Set initial focus to scanning tab and load initial data
    root.after(100, lambda: app.notebook.select(0))  # Select scanning tab (index 0)
    root.after(200, lambda: app.refresh_scanning_history())  # Load initial history data
    root.after(300, lambda: app.load_today_summary())  # Load initial summary data
    
    root.mainloop()

if __name__ == "__main__":
    main()